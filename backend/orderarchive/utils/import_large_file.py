# backend/orderarchive/utils/import_large_file.py
from openpyxl import load_workbook
from orderarchive.models import OrderDetail
from datetime import datetime

def preserve_text(value):
    """
    Gelen değeri metin olarak döndürür. None değerleri boş stringe çevirir.
    """
    try:
        if value is None:
            return ""
        return str(value).strip()
    except Exception:
        return ""

def convert_to_int_or_none(value):
    """
    Sayısal bir değeri integer olarak döndürür, boşsa None döndürür.
    """
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(float(str(value).replace(".", "").replace(",", ".")))
    except (ValueError, TypeError):
        return None

def convert_to_date(value):
    """
    Tarih değerlerini (örn: '31.12.2023') Python date nesnesine çevirir.
    """
    try:
        if isinstance(value, datetime):
            return value.date()
        return datetime.strptime(value, "%d.%m.%Y").date()
    except (ValueError, TypeError):
        return None

def import_large_file(file_path, chunk_size=10000):
    """
    Büyük Excel dosyalarını parça parça veritabanına aktarır.
    """
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

    rows_to_save = []
    errors = []
    row_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        row_count += 1

        try:
            rows_to_save.append(OrderDetail(
                seller=preserve_text(row_data.get("Satici")),
                order_number=preserve_text(row_data.get("SipNo")),
                order_date=convert_to_date(row_data.get("SipTarih")),
                year=convert_to_int_or_none(row_data.get("Yil")),  # Yıl için None döndürüyoruz
                month=convert_to_int_or_none(row_data.get("Ay")),  # Ay için None döndürüyoruz
                delivery_date=convert_to_date(row_data.get("TeslimTarih")),
                country=preserve_text(row_data.get("Ulke")),
                city=preserve_text(row_data.get("Sehir")),
                customer_code=preserve_text(row_data.get("MusteriKod")),
                customer_name=preserve_text(row_data.get("MusteriAd")),
                document_description=preserve_text(row_data.get("BelgeAciklama")),
                color_code=preserve_text(row_data.get("RenkKod")),
                detail_description=preserve_text(row_data.get("DetayAciklama1-2-3")),
                line_number=convert_to_int_or_none(row_data.get("SiraNo")),
                item_code=preserve_text(row_data.get("KalemKod")),
                item_description=preserve_text(row_data.get("KalemTanim")),
                quantity=preserve_text(row_data.get("Miktar")),
                unit_price=preserve_text(row_data.get("BirimFiyat")),
                vat_percentage=preserve_text(row_data.get("KdvYuzde")),
                vat_amount=preserve_text(row_data.get("KdvTutar")),
                discount_rate=preserve_text(row_data.get("IskOran")),
                discount_amount=preserve_text(row_data.get("IsktoluTutar")),
                currency=preserve_text(row_data.get("Doviz")),
                exchange_rate=preserve_text(row_data.get("Kur")),
                currency_price=preserve_text(row_data.get("DovizFiyat")),
                currency_movement_amount=preserve_text(row_data.get("DovizHareketTutar")),
            ))
        except Exception as e:
            errors.append(f"Satır {row_count}: {e} - {row_data}")
            continue

        if len(rows_to_save) >= chunk_size:
            OrderDetail.objects.bulk_create(rows_to_save)
            print(f"{len(rows_to_save)} satır yüklendi.")
            rows_to_save = []

    if rows_to_save:
        OrderDetail.objects.bulk_create(rows_to_save)
        print(f"{len(rows_to_save)} satır yüklendi.")

    if errors:
        error_log_file = "/var/www/sapb1reportsv2/backend/logs/error_log.txt"
        with open(error_log_file, "a") as f:
            for error in errors:
                f.write(error + "\n")
        print(f"{len(errors)} hata loglandı. Ayrıntılar: {error_log_file}")

    print(f"Toplam {row_count} satır işlendi.")

