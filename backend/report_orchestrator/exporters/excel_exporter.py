# backend/report_orchestrator/exporters/excel_exporter.py

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.conf import settings

EXPORT_DIR = os.path.join(settings.BASE_DIR, "data/exports/excel")


def generate_excel_from_result(api_name: str, result_json: dict) -> str:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    """
    result_json içeriğine göre Excel dosyası oluşturur ve belirtilen klasöre kaydeder.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"

    # Font ayarları
    bold_font = Font(size=16, bold=True)
    normal_font = Font(size=14)
    align_center = Alignment(horizontal="center", vertical="center")

    # Başlık ekle
    report_title = result_json.get("report_title", api_name)
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = report_title
    cell.font = bold_font
    cell.alignment = align_center

    # Tarih
    ws.merge_cells("A2:D2")
    cell = ws["A2"]
    cell.value = f"Rapor Tarihi: {result_json.get('report_date', datetime.today().strftime('%Y-%m-%d'))}"
    cell.font = normal_font
    cell.alignment = align_center

    # Kolon başlıkları
    columns = result_json.get("columns") or list(result_json.get("data", [{}])[0].keys())
    for col_num, col in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_num, value=col)
        cell.font = bold_font
        cell.alignment = align_center

    # Veriler
    for row_idx, row in enumerate(result_json.get("data", []), start=5):
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col, "") if isinstance(row, dict) else row[col_idx-1])
            cell.font = normal_font

    # Alt toplamlar
    current_row = row_idx + 1
    if "ara_toplam" in result_json:
        ws.cell(row=current_row, column=len(columns)-1, value="Ara Toplam:").font = bold_font
        ws.cell(row=current_row, column=len(columns), value=result_json["ara_toplam"]).font = bold_font
        current_row += 1
    if "kumulatif_toplam" in result_json:
        ws.cell(row=current_row, column=len(columns)-1, value="Kümülatif:").font = bold_font
        ws.cell(row=current_row, column=len(columns), value=result_json["kumulatif_toplam"]).font = bold_font

    # Dosya kaydetme
    filename = f"{api_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return filepath

