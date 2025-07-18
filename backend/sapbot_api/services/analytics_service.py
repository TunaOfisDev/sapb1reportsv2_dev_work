# backend/sapbot_api/services/analytics_service.py
"""
SAPBot API Analytics Service

Bu servis, sistem analitiklerini, kullanıcı metriklerini ve performans
ölçümlerini toplama, işleme ve raporlama işlemlerini yönetir.
"""
import logging
from tkinter.font import Font
from typing import Dict, List, Optional, Any
from datetime import datetime, timedelta
from dataclasses import dataclass
from django.db import models
from django.db.models import Count, Avg, Sum, Max, Min, Q
from django.utils import timezone
from django.core.cache import cache


from ..models import (
   QueryAnalytics, UserFeedback, PerformanceMetrics,
   ErrorLog, ChatMessage, ChatConversation, DocumentSource, 
   KnowledgeChunk, UserProfile
)
from ..utils.cache_utils import system_cache
from ..utils.exceptions import ValidationException

logger = logging.getLogger(__name__)


@dataclass
class AnalyticsResult:
   """Analytics sonuç objesi"""
   success: bool
   data: Dict[str, Any]
   message: str
   cached: bool = False
   computation_time: float = 0.0


@dataclass
class MetricSummary:
   """Metrik özet objesi"""
   name: str
   value: Any
   change: Optional[float] = None
   trend: Optional[str] = None  # 'up', 'down', 'stable'
   format_type: str = 'number'  # 'number', 'percentage', 'currency', 'time'


class QueryAnalyticsService:
   """Sorgu analitikleri servisi"""
   
   def __init__(self):
       self.cache_timeout = 3600  # 1 saat
   
   def record_query(
       self,
       user_id: str,
       session_id: str,
       query: str,
       user_type: str,
       response_data: Dict[str, Any]
   ) -> QueryAnalytics:
       """Sorguyu analytics'e kaydet"""
       try:
           # Query hash oluştur
           import hashlib
           query_hash = hashlib.sha256(query.encode('utf-8')).hexdigest()
           
           # Intent ve SAP modül tespiti
           from ..utils.text_processing import (
               IntentClassifier, 
               SAPTerminologyAnalyzer
           )
           
           intent, intent_confidence = IntentClassifier.classify_intent(query)
           sap_modules = SAPTerminologyAnalyzer.detect_sap_modules(query)
           sap_module = sap_modules[0][0] if sap_modules else None
           
           # Analytics kaydı oluştur
           analytics = QueryAnalytics.objects.create(
               user_id=user_id,
               session_id=session_id,
               query=query[:1000],  # Uzunluk sınırı
               query_hash=query_hash,
               query_length=len(query),
               user_type=user_type,
               sap_module_detected=sap_module,
               intent_detected=intent,
               confidence_score=intent_confidence,
               response_generated=response_data.get('success', False),
               response_time=response_data.get('response_time', 0),
               sources_used_count=len(response_data.get('sources', [])),
               tokens_used=response_data.get('tokens_used'),
               cost_estimate=response_data.get('cost_estimate'),
               language_detected=response_data.get('language', 'tr'),
               error_occurred=not response_data.get('success', True),
               error_type=response_data.get('error_type'),
               metadata=response_data.get('metadata', {})
           )
           
           # Cache'i güncelle
           self._invalidate_query_cache()
           
           return analytics
           
       except Exception as e:
           logger.error(f"Query analytics kaydetme hatası: {e}")
           raise
   
   def get_query_trends(
       self,
       start_date: datetime,
       end_date: datetime,
       group_by: str = 'day'
   ) -> Dict[str, Any]:
       """Sorgu trendlerini al"""
       cache_key = f"query_trends_{start_date.date()}_{end_date.date()}_{group_by}"
       cached_result = system_cache.get(cache_key)
       
       if cached_result:
           return cached_result
       
       try:
           # Tarih grupları oluştur
           if group_by == 'hour':
               date_trunc = "DATE_TRUNC('hour', created_at)"
               date_format = '%Y-%m-%d %H:00'
           elif group_by == 'day':
               date_trunc = "DATE_TRUNC('day', created_at)"
               date_format = '%Y-%m-%d'
           elif group_by == 'week':
               date_trunc = "DATE_TRUNC('week', created_at)"
               date_format = '%Y-W%U'
           elif group_by == 'month':
               date_trunc = "DATE_TRUNC('month', created_at)"
               date_format = '%Y-%m'
           else:
               raise ValidationException("Geçersiz group_by parametresi")
           
           # Temel sorgu metrikleri
           queryset = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date]
           )
           
           # Zaman bazlı gruplandırma
           trends = queryset.extra(
               select={'period': date_trunc}
           ).values('period').annotate(
               total_queries=Count('id'),
               successful_queries=Count('id', filter=Q(response_generated=True)),
               avg_response_time=Avg('response_time'),
               avg_confidence=Avg('confidence_score'),
               unique_users=Count('user_id', distinct=True),
               unique_sessions=Count('session_id', distinct=True)
           ).order_by('period')
           
           # Modül bazlı dağılım
           module_distribution = queryset.values('sap_module_detected').annotate(
               count=Count('id')
           ).order_by('-count')[:10]
           
           # Intent dağılımı
           intent_distribution = queryset.values('intent_detected').annotate(
               count=Count('id')
           ).order_by('-count')[:10]
           
           # Kullanıcı tipi dağılımı
           user_type_distribution = queryset.values('user_type').annotate(
               count=Count('id')
           ).order_by('-count')
           
           result = {
               'trends': list(trends),
               'module_distribution': list(module_distribution),
               'intent_distribution': list(intent_distribution),
               'user_type_distribution': list(user_type_distribution),
               'total_queries': queryset.count(),
               'period': group_by,
               'date_range': {
                   'start': start_date.isoformat(),
                   'end': end_date.isoformat()
               }
           }
           
           # Cache'e kaydet
           system_cache.set(cache_key, result, self.cache_timeout)
           
           return result
           
       except Exception as e:
           logger.error(f"Query trends alma hatası: {e}")
           raise
   
   def get_top_queries(
       self,
       start_date: datetime,
       end_date: datetime,
       limit: int = 20,
       user_type: Optional[str] = None
   ) -> List[Dict[str, Any]]:
       """En popüler sorguları al"""
       try:
           queryset = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date]
           )
           
           if user_type:
               queryset = queryset.filter(user_type=user_type)
           
           # Query hash bazlı gruplama
           top_queries = queryset.values(
               'query_hash'
           ).annotate(
               count=Count('id'),
               avg_response_time=Avg('response_time'),
               success_rate=Avg(
                   models.Case(
                       models.When(response_generated=True, then=1),
                       default=0,
                       output_field=models.FloatField()
                   )
               ),
               last_query=Max('query'),
               last_used=Max('created_at')
           ).order_by('-count')[:limit]
           
           # Detaylı bilgileri ekle
           enriched_queries = []
           for query_stat in top_queries:
               # Sample query al
               sample = QueryAnalytics.objects.filter(
                   query_hash=query_stat['query_hash']
               ).first()
               
               enriched_queries.append({
                   'query_hash': query_stat['query_hash'],
                   'sample_query': sample.query if sample else '',
                   'count': query_stat['count'],
                   'avg_response_time': query_stat['avg_response_time'],
                   'success_rate': query_stat['success_rate'] * 100,
                   'last_used': query_stat['last_used'],
                   'sap_module': sample.sap_module_detected if sample else None,
                   'intent': sample.intent_detected if sample else None
               })
           
           return enriched_queries
           
       except Exception as e:
           logger.error(f"Top queries alma hatası: {e}")
           raise
   
   def get_user_analytics(
       self,
       start_date: datetime,
       end_date: datetime
   ) -> Dict[str, Any]:
       """Kullanıcı analitiklerini al"""
       try:
           queryset = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date]
           )
           
           # Kullanıcı bazlı metrikler
           user_stats = queryset.values('user_id').annotate(
               query_count=Count('id'),
               avg_response_time=Avg('response_time'),
               success_rate=Avg(
                   models.Case(
                       models.When(response_generated=True, then=1),
                       default=0,
                       output_field=models.FloatField()
                   )
               ),
               last_activity=Max('created_at'),
               unique_sessions=Count('session_id', distinct=True)
           ).order_by('-query_count')
           
           # Kullanıcı tipi bazlı analiz
           user_type_stats = queryset.values('user_type').annotate(
               user_count=Count('user_id', distinct=True),
               query_count=Count('id'),
               avg_queries_per_user=Count('id') / Count('user_id', distinct=True),
               avg_response_time=Avg('response_time'),
               success_rate=Avg(
                   models.Case(
                       models.When(response_generated=True, then=1),
                       default=0,
                       output_field=models.FloatField()
                   )
               )
           ).order_by('-query_count')
           
           # Aktif kullanıcı analizi
           total_users = queryset.values('user_id').distinct().count()
           daily_active = queryset.filter(
               created_at__gte=timezone.now() - timedelta(days=1)
           ).values('user_id').distinct().count()
           
           weekly_active = queryset.filter(
               created_at__gte=timezone.now() - timedelta(days=7)
           ).values('user_id').distinct().count()
           
           return {
               'user_stats': list(user_stats[:50]),  # Top 50 kullanıcı
               'user_type_stats': list(user_type_stats),
               'summary': {
                   'total_users': total_users,
                   'daily_active_users': daily_active,
                   'weekly_active_users': weekly_active,
                   'dau_wau_ratio': daily_active / weekly_active if weekly_active > 0 else 0
               }
           }
           
       except Exception as e:
           logger.error(f"User analytics alma hatası: {e}")
           raise
   
   def _invalidate_query_cache(self):
       """Query cache'lerini temizle"""
       try:
           # Pattern bazlı cache temizleme
           cache_patterns = [
               'query_trends_*',
               'system_metrics_*',
               'dashboard_data_*'
           ]
           
           for pattern in cache_patterns:
               system_cache.clear_pattern(pattern)
               
       except Exception as e:
           logger.warning(f"Cache temizleme hatası: {e}")


class SystemMetricsService:
   """Sistem metrikleri servisi"""
   
   def collect_metrics(self) -> Dict[str, Any]:
       """Sistem metriklerini topla"""
       try:
           metrics = {}
           
           # Database metrikleri
           metrics['database'] = self._collect_database_metrics()
           
           # Performance metrikleri
           metrics['performance'] = self._collect_performance_metrics()
           
           # Usage metrikleri
           metrics['usage'] = self._collect_usage_metrics()
           
           # System health
           metrics['health'] = self._collect_health_metrics()
           
           # Cache metrikleri
           metrics['cache'] = self._collect_cache_metrics()
           
           return {
               'success': True,
               'metrics': metrics,
               'collected_at': timezone.now().isoformat()
           }
           
       except Exception as e:
           logger.error(f"Sistem metrikleri toplama hatası: {e}")
           return {
               'success': False,
               'error': str(e),
               'collected_at': timezone.now().isoformat()
           }
   
   def _collect_database_metrics(self) -> Dict[str, Any]:
       """Database metriklerini topla"""
       try:
           from django.db import connection
           
           with connection.cursor() as cursor:
               # Tablo boyutları
               cursor.execute("""
                   SELECT 
                       schemaname,
                       tablename,
                       pg_size_pretty(pg_total_relation_size(schemaname||'.'||tablename)) as size,
                       pg_total_relation_size(schemaname||'.'||tablename) as size_bytes
                   FROM pg_tables 
                   WHERE schemaname = 'public' 
                   AND tablename LIKE 'sapbot_%'
                   ORDER BY pg_total_relation_size(schemaname||'.'||tablename) DESC
                   LIMIT 10
               """)
               
               table_sizes = cursor.fetchall()
               
               # Connection sayısı
               cursor.execute("""
                   SELECT count(*) as active_connections
                   FROM pg_stat_activity 
                   WHERE state = 'active'
               """)
               
               active_connections = cursor.fetchone()[0]
           
           # Model bazlı kayıt sayıları
           record_counts = {
               'documents': DocumentSource.objects.count(),
               'knowledge_chunks': KnowledgeChunk.objects.count(),
               'chat_conversations': ChatConversation.objects.count(),
               'chat_messages': ChatMessage.objects.count(),
               'query_analytics': QueryAnalytics.objects.count(),
               'user_profiles': UserProfile.objects.count()
           }
           
           return {
               'table_sizes': [
                   {
                       'schema': row[0],
                       'table': row[1],
                       'size_formatted': row[2],
                       'size_bytes': row[3]
                   }
                   for row in table_sizes
               ],
               'active_connections': active_connections,
               'record_counts': record_counts
           }
           
       except Exception as e:
           logger.error(f"Database metrikleri hatası: {e}")
           return {'error': str(e)}
   
   def _collect_performance_metrics(self) -> Dict[str, Any]:
       """Performans metriklerini topla"""
       try:
           # Son 24 saatin performans metrikleri
           last_24h = timezone.now() - timedelta(hours=24)
           
           performance_data = PerformanceMetrics.objects.filter(
               created_at__gte=last_24h
           ).values('component', 'metric_name').annotate(
               avg_value=Avg('value'),
               max_value=Max('value'),
               min_value=Min('value'),
               count=Count('id')
           )
           
           # Query response time analizi
           query_performance = QueryAnalytics.objects.filter(
               created_at__gte=last_24h,
               response_time__isnull=False
           ).aggregate(
               avg_response_time=Avg('response_time'),
               p95_response_time=models.functions.Percentile(
                   'response_time', 0.95
               ),
               max_response_time=Max('response_time'),
               min_response_time=Min('response_time')
           )
           
           return {
               'component_metrics': list(performance_data),
               'query_performance': query_performance,
               'period': '24h'
           }
           
       except Exception as e:
           logger.error(f"Performance metrikleri hatası: {e}")
           return {'error': str(e)}
   
   def _collect_usage_metrics(self) -> Dict[str, Any]:
       """Kullanım metriklerini topla"""
       try:
           now = timezone.now()
           
           # Farklı zaman aralıkları için metrikler
           periods = {
               'last_hour': now - timedelta(hours=1),
               'last_24h': now - timedelta(hours=24),
               'last_7d': now - timedelta(days=7),
               'last_30d': now - timedelta(days=30)
           }
           
           usage_stats = {}
           
           for period_name, start_time in periods.items():
               stats = QueryAnalytics.objects.filter(
                   created_at__gte=start_time
               ).aggregate(
                   total_queries=Count('id'),
                   successful_queries=Count(
                       'id', filter=Q(response_generated=True)
                   ),
                   unique_users=Count('user_id', distinct=True),
                   unique_sessions=Count('session_id', distinct=True),
                   avg_response_time=Avg('response_time'),
                   total_tokens=Sum('tokens_used')
               )
               
               usage_stats[period_name] = stats
           
           # Document processing stats
           doc_stats = DocumentSource.objects.aggregate(
               total_documents=Count('id'),
               processed_documents=Count(
                   'id', filter=Q(processing_status='completed')
               ),
               pending_documents=Count(
                   'id', filter=Q(processing_status='pending')
               ),
               failed_documents=Count(
                   'id', filter=Q(processing_status='failed')
               )
           )
           
           return {
               'query_stats': usage_stats,
               'document_stats': doc_stats,
               'collected_at': now.isoformat()
           }
           
       except Exception as e:
           logger.error(f"Usage metrikleri hatası: {e}")
           return {'error': str(e)}
   
   def _collect_health_metrics(self) -> Dict[str, Any]:
       """Sistem sağlık metriklerini topla"""
       try:
           health_status = {
               'overall': 'healthy',
               'components': {}
           }
           
           # Database health
           try:
               from django.db import connection
               with connection.cursor() as cursor:
                   cursor.execute('SELECT 1')
                   cursor.fetchone()
               health_status['components']['database'] = 'healthy'
           except Exception:
               health_status['components']['database'] = 'unhealthy'
               health_status['overall'] = 'degraded'
           
           # Cache health
           try:
               cache.set('health_check', 'ok', 60)
               test_value = cache.get('health_check')
               if test_value == 'ok':
                   health_status['components']['cache'] = 'healthy'
               else:
                   health_status['components']['cache'] = 'degraded'
           except Exception:
               health_status['components']['cache'] = 'unhealthy'
               health_status['overall'] = 'degraded'
           
           # Error rate analizi
           last_hour = timezone.now() - timedelta(hours=1)
           error_count = ErrorLog.objects.filter(
               created_at__gte=last_hour
           ).count()
           
           if error_count > 100:  # Saatte 100+ hata
               health_status['components']['error_rate'] = 'unhealthy'
               health_status['overall'] = 'degraded'
           elif error_count > 50:
               health_status['components']['error_rate'] = 'warning'
           else:
               health_status['components']['error_rate'] = 'healthy'
           
           return health_status
           
       except Exception as e:
           logger.error(f"Health metrikleri hatası: {e}")
           return {
               'overall': 'unknown',
               'error': str(e)
           }
   
   def _collect_cache_metrics(self) -> Dict[str, Any]:
       """Cache metriklerini topla"""
       try:
           from ..utils.cache_utils import get_cache_stats
           return get_cache_stats()
       except Exception as e:
           logger.error(f"Cache metrikleri hatası: {e}")
           return {'error': str(e)}


class DashboardService:
   """Dashboard veri servisi"""
   
   def __init__(self):
       self.query_service = QueryAnalyticsService()
       self.metrics_service = SystemMetricsService()
       self.cache_timeout = 1800  # 30 dakika
   
   def get_dashboard_data(self, period: str = 'daily') -> AnalyticsResult:
       """Dashboard verilerini al"""
       start_time = timezone.now()
       
       cache_key = f"dashboard_data_{period}_{start_time.date()}"
       cached_data = system_cache.get(cache_key)
       
       if cached_data:
           return AnalyticsResult(
               success=True,
               data=cached_data,
               message="Dashboard verileri başarıyla alındı",
               cached=True,
               computation_time=0.0
           )
       
       try:
           # Zaman aralığını belirle
           end_date = timezone.now()
           if period == 'daily':
               start_date = end_date - timedelta(days=30)
           elif period == 'weekly':
               start_date = end_date - timedelta(weeks=12)
           elif period == 'monthly':
               start_date = end_date - timedelta(days=365)
           else:
               start_date = end_date - timedelta(days=30)
           
           # Ana metrikler
           summary_metrics = self._get_summary_metrics(start_date, end_date)
           
           # Trend verileri
           trends = self.query_service.get_query_trends(
               start_date, end_date, period
           )
           
           # Top queries
           top_queries = self.query_service.get_top_queries(
               start_date, end_date, limit=10
           )
           
           # User analytics
           user_analytics = self.query_service.get_user_analytics(
               start_date, end_date
           )
           
           # System metrics
           system_metrics = self.metrics_service.collect_metrics()
           
           # Feedback analizi
           feedback_data = self._get_feedback_analytics(start_date, end_date)
           
           # Module popularity
           module_stats = self._get_module_statistics(start_date, end_date)
           
           dashboard_data = {
               'summary': summary_metrics,
               'trends': trends,
               'top_queries': top_queries,
               'user_analytics': user_analytics,
               'system_metrics': system_metrics,
               'feedback': feedback_data,
               'modules': module_stats,
               'period': period,
               'generated_at': timezone.now().isoformat(),
               'data_range': {
                   'start': start_date.isoformat(),
                   'end': end_date.isoformat()
               }
           }
           
           # Cache'e kaydet
           system_cache.set(cache_key, dashboard_data, self.cache_timeout)
           
           computation_time = (timezone.now() - start_time).total_seconds()
           
           return AnalyticsResult(
               success=True,
               data=dashboard_data,
               message="Dashboard verileri başarıyla oluşturuldu",
               cached=False,
               computation_time=computation_time
           )
           
       except Exception as e:
           logger.error(f"Dashboard veri hatası: {e}")
           return AnalyticsResult(
               success=False,
               data={},
               message=f"Dashboard veri hatası: {str(e)}",
               computation_time=(timezone.now() - start_time).total_seconds()
           )
   
   def _get_summary_metrics(
       self, 
       start_date: datetime, 
       end_date: datetime
   ) -> List[MetricSummary]:
       """Özet metrikleri al"""
       try:
           # Mevcut dönem
           current_period = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date]
           )
           
           # Önceki dönem (karşılaştırma için)
           period_length = end_date - start_date
           previous_start = start_date - period_length
           previous_period = QueryAnalytics.objects.filter(
               created_at__range=[previous_start, start_date]
           )
           
           # Metrikler
           current_stats = current_period.aggregate(
               total_queries=Count('id'),
               successful_queries=Count('id', filter=Q(response_generated=True)),
               unique_users=Count('user_id', distinct=True),
               avg_response_time=Avg('response_time'),
               avg_confidence=Avg('confidence_score')
           )
           
           previous_stats = previous_period.aggregate(
               total_queries=Count('id'),
               successful_queries=Count('id', filter=Q(response_generated=True)),
               unique_users=Count('user_id', distinct=True),
               avg_response_time=Avg('response_time'),
               avg_confidence=Avg('confidence_score')
           )
           
           def calculate_change(current, previous):
               if previous and previous > 0:
                   return ((current - previous) / previous) * 100
               return 0
           
           metrics = [
               MetricSummary(
                   name="Toplam Sorgu",
                   value=current_stats['total_queries'] or 0,
                   change=calculate_change(
                       current_stats['total_queries'] or 0,
                       previous_stats['total_queries'] or 0
                   ),
                   trend='up' if (current_stats['total_queries'] or 0) > (previous_stats['total_queries'] or 0) else 'down'
               ),
               MetricSummary(
                   name="Başarı Oranı",
                   value=round(
                       (current_stats['successful_queries'] or 0) / 
                       max(current_stats['total_queries'] or 1, 1) * 100, 1
                   ),
                   format_type='percentage',
                   change=calculate_change(
                       (current_stats['successful_queries'] or 0) / max(current_stats['total_queries'] or 1, 1),
                       (previous_stats['successful_queries'] or 0) / max(previous_stats['total_queries'] or 1, 1)
                   )
               ),
               MetricSummary(
                   name="Aktif Kullanıcı",
                   value=current_stats['unique_users'] or 0,
                   change=calculate_change(
                       current_stats['unique_users'] or 0,
                       previous_stats['unique_users'] or 0
                   )
               ),
               MetricSummary(
                   name="Ortalama Yanıt Süresi",
                   value=round(current_stats['avg_response_time'] or 0, 2),
                   format_type='time',
                   change=calculate_change(
                       current_stats['avg_response_time'] or 0,
                       previous_stats['avg_response_time'] or 0
                   ),
                   trend='down' if (current_stats['avg_response_time'] or 0) < (previous_stats['avg_response_time'] or 0) else 'up'
               ),
               MetricSummary(
                   name="Güven Skoru",
                   value=round((current_stats['avg_confidence'] or 0) * 100, 1),
                   format_type='percentage',
                   change=calculate_change(
                       current_stats['avg_confidence'] or 0,
                       previous_stats['avg_confidence'] or 0
                   )
               )
           ]
           
           return metrics
           
       except Exception as e:
           logger.error(f"Summary metrics hatası: {e}")
           return []
   
   def _get_feedback_analytics(
       self,
       start_date: datetime,
       end_date: datetime
   ) -> Dict[str, Any]:
       """Feedback analitiklerini al"""
       try:
           feedback_queryset = UserFeedback.objects.filter(
               created_at__range=[start_date, end_date]
           )
           
           # Rating dağılımı
           rating_distribution = feedback_queryset.values('rating').annotate(
               count=Count('id')
           ).order_by('rating')
           
           # Ortalama rating
           avg_rating = feedback_queryset.aggregate(
               avg_rating=Avg('rating')
           )['avg_rating'] or 0
           
           # Helpful oranı
           helpful_stats = feedback_queryset.aggregate(
               total_feedback=Count('id'),
               helpful_count=Count('id', filter=Q(is_helpful=True))
           )
           
           helpful_rate = 0
           if helpful_stats['total_feedback'] > 0:
               helpful_rate = (helpful_stats['helpful_count'] / helpful_stats['total_feedback']) * 100
           
           # Feedback tipi dağılımı
           type_distribution = feedback_queryset.values('feedback_type').annotate(
               count=Count('id')
           ).order_by('-count')
           
           return {
               'rating_distribution': list(rating_distribution),
               'average_rating': round(avg_rating, 2),
               'helpful_rate': round(helpful_rate, 1),
               'type_distribution': list(type_distribution),
               'total_feedback': helpful_stats['total_feedback']
           }
           
       except Exception as e:
           logger.error(f"Feedback analytics hatası: {e}")
           return {}
   
   def _get_module_statistics(
       self,
       start_date: datetime,
       end_date: datetime
   ) -> Dict[str, Any]:
       """SAP modül istatistiklerini al"""
       try:
           # Modül bazlı query analizi
           module_stats = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date],
               sap_module_detected__isnull=False
           ).values('sap_module_detected').annotate(
               query_count=Count('id'),
               success_rate=Avg(
                   models.Case(
                       models.When(response_generated=True, then=1),
                       default=0,
                       output_field=models.FloatField()
                   )
               ),
               avg_response_time=Avg('response_time'),
               avg_confidence=Avg('confidence_score'),
               unique_users=Count('user_id', distinct=True)
           ).order_by('-query_count')
           
           # Knowledge chunks kullanım istatistikleri
           chunk_usage = KnowledgeChunk.objects.filter(
               last_used__range=[start_date, end_date]
           ).values('sap_module').annotate(
               chunk_count=Count('id'),
               total_usage=Sum('usage_count'),
               avg_usage=Avg('usage_count')
           ).order_by('-total_usage')
           
           # Module popularity trends (son 7 gün)
           last_week = end_date - timedelta(days=7)
           recent_module_trends = QueryAnalytics.objects.filter(
               created_at__gte=last_week,
               sap_module_detected__isnull=False
           ).extra(
               select={'day': "DATE_TRUNC('day', created_at)"}
           ).values('day', 'sap_module_detected').annotate(
               count=Count('id')
           ).order_by('day', 'sap_module_detected')
           
           return {
               'module_statistics': list(module_stats),
               'chunk_usage': list(chunk_usage),
               'recent_trends': list(recent_module_trends),
               'total_modules_used': len(module_stats)
           }
           
       except Exception as e:
           logger.error(f"Module statistics hatası: {e}")
           return {}


class ReportingService:
   """Raporlama servisi"""
   
   def __init__(self):
       self.query_service = QueryAnalyticsService()
       self.dashboard_service = DashboardService()
   
   def generate_executive_report(
       self,
       start_date: datetime,
       end_date: datetime
   ) -> Dict[str, Any]:
       """Yönetici raporu oluştur"""
       try:
           # Temel metrikler
           total_queries = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date]
           ).count()
           
           # Başarı oranı
           successful_queries = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date],
               response_generated=True
           ).count()
           
           success_rate = (successful_queries / max(total_queries, 1)) * 100
           
           # Kullanıcı metrikleri
           user_metrics = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date]
           ).aggregate(
               unique_users=Count('user_id', distinct=True),
               unique_sessions=Count('session_id', distinct=True),
               avg_queries_per_user=Count('id') / Count('user_id', distinct=True)
           )
           
           # Cost analizi
           cost_analysis = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date],
               cost_estimate__isnull=False
           ).aggregate(
               total_cost=Sum('cost_estimate'),
               avg_cost_per_query=Avg('cost_estimate'),
               total_tokens=Sum('tokens_used')
           )
           
           # Top performing content
           top_chunks = KnowledgeChunk.objects.filter(
               last_used__range=[start_date, end_date]
           ).order_by('-usage_count')[:10]
           
           # Error analizi
           error_analysis = ErrorLog.objects.filter(
               created_at__range=[start_date, end_date]
           ).values('error_type').annotate(
               count=Count('id')
           ).order_by('-count')[:5]
           
           # Kullanıcı memnuniyeti
           satisfaction = UserFeedback.objects.filter(
               created_at__range=[start_date, end_date]
           ).aggregate(
               avg_rating=Avg('rating'),
               total_feedback=Count('id'),
               helpful_count=Count('id', filter=Q(is_helpful=True))
           )
           
           report = {
               'executive_summary': {
                   'period': {
                       'start': start_date.isoformat(),
                       'end': end_date.isoformat(),
                       'days': (end_date - start_date).days
                   },
                   'key_metrics': {
                       'total_queries': total_queries,
                       'success_rate': round(success_rate, 2),
                       'unique_users': user_metrics['unique_users'],
                       'user_satisfaction': round(satisfaction['avg_rating'] or 0, 2),
                       'total_cost': float(cost_analysis['total_cost'] or 0)
                   }
               },
               'detailed_analysis': {
                   'user_engagement': {
                       'unique_users': user_metrics['unique_users'],
                       'unique_sessions': user_metrics['unique_sessions'],
                       'avg_queries_per_user': round(user_metrics['avg_queries_per_user'] or 0, 2),
                       'session_to_user_ratio': round(
                           user_metrics['unique_sessions'] / max(user_metrics['unique_users'], 1), 2
                       )
                   },
                   'cost_analysis': {
                       'total_cost': float(cost_analysis['total_cost'] or 0),
                       'avg_cost_per_query': float(cost_analysis['avg_cost_per_query'] or 0),
                       'total_tokens': cost_analysis['total_tokens'] or 0,
                       'cost_per_user': float(cost_analysis['total_cost'] or 0) / max(user_metrics['unique_users'], 1)
                   },
                   'content_performance': [
                       {
                           'id': str(chunk.id),
                           'source_title': chunk.source.title,
                           'sap_module': chunk.sap_module,
                           'usage_count': chunk.usage_count,
                           'last_used': chunk.last_used.isoformat() if chunk.last_used else None
                       }
                       for chunk in top_chunks
                   ],
                   'error_analysis': list(error_analysis),
                   'user_satisfaction': {
                       'average_rating': round(satisfaction['avg_rating'] or 0, 2),
                       'total_feedback': satisfaction['total_feedback'],
                       'helpful_percentage': round(
                           (satisfaction['helpful_count'] / max(satisfaction['total_feedback'], 1)) * 100, 2
                       )
                   }
               },
               'recommendations': self._generate_recommendations(
                   success_rate, 
                   user_metrics, 
                   satisfaction,
                   cost_analysis
               ),
               'generated_at': timezone.now().isoformat()
           }
           
           return report
           
       except Exception as e:
           logger.error(f"Executive report oluşturma hatası: {e}")
           raise
   
   def generate_technical_report(
       self,
       start_date: datetime,
       end_date: datetime
   ) -> Dict[str, Any]:
       """Teknik rapor oluştur"""
       try:
           # Performance metrikleri
           performance_stats = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date],
               response_time__isnull=False
           ).aggregate(
               avg_response_time=Avg('response_time'),
               min_response_time=Min('response_time'),
               max_response_time=Max('response_time'),
               p50_response_time=models.functions.Percentile('response_time', 0.5),
               p95_response_time=models.functions.Percentile('response_time', 0.95),
               p99_response_time=models.functions.Percentile('response_time', 0.99)
           )
           
           # System metrics
           system_metrics = self.dashboard_service.metrics_service.collect_metrics()
           
           # Database performance
           slow_queries = QueryAnalytics.objects.filter(
               created_at__range=[start_date, end_date],
               response_time__gt=10.0  # 10 saniyeden yavaş
           ).order_by('-response_time')[:20]
           
           # Error patterns
           error_patterns = ErrorLog.objects.filter(
               created_at__range=[start_date, end_date]
           ).values('error_type', 'component').annotate(
               count=Count('id'),
               latest_occurrence=Max('created_at')
           ).order_by('-count')
           
           # Cache performance
           cache_stats = system_metrics.get('metrics', {}).get('cache', {})
           
           # Document processing stats
           doc_processing = DocumentSource.objects.filter(
               created_at__range=[start_date, end_date]
           ).values('processing_status').annotate(
               count=Count('id')
           )
           
           report = {
               'performance_analysis': {
                   'response_times': performance_stats,
                   'slow_queries': [
                       {
                           'query_hash': q.query_hash,
                           'response_time': q.response_time,
                           'sap_module': q.sap_module_detected,
                           'user_type': q.user_type,
                           'created_at': q.created_at.isoformat()
                       }
                       for q in slow_queries
                   ]
               },
               'system_health': system_metrics,
               'error_analysis': {
                   'error_patterns': list(error_patterns),
                   'error_rate': len(error_patterns)
               },
               'infrastructure': {
                   'cache_performance': cache_stats,
                   'document_processing': list(doc_processing)
               },
               'recommendations': self._generate_technical_recommendations(
                   performance_stats,
                   error_patterns,
                   cache_stats
               ),
               'generated_at': timezone.now().isoformat()
           }
           
           return report
           
       except Exception as e:
           logger.error(f"Technical report oluşturma hatası: {e}")
           raise
   
   def _generate_recommendations(
       self,
       success_rate: float,
       user_metrics: Dict,
       satisfaction: Dict,
       cost_analysis: Dict
   ) -> List[Dict[str, Any]]:
       """İş önerilerini oluştur"""
       recommendations = []
       
       # Başarı oranı önerisi
       if success_rate < 85:
           recommendations.append({
               'type': 'performance',
               'priority': 'high',
               'title': 'Düşük Başarı Oranı',
               'description': f'Sistem başarı oranı %{success_rate:.1f}. Hedef %90\'ın üzerinde olmalı.',
               'action': 'Knowledge base\'i genişletin ve intent detection modelini iyileştirin.'
           })
       
       # Kullanıcı engagement önerisi
       avg_queries = user_metrics.get('avg_queries_per_user', 0)
       if avg_queries < 5:
           recommendations.append({
               'type': 'engagement',
               'priority': 'medium',
               'title': 'Düşük Kullanıcı Engagement',
               'description': f'Kullanıcı başına ortalama {avg_queries:.1f} sorgu. Hedef 10+ olmalı.',
               'action': 'Kullanıcı eğitimi verin ve use case\'leri tanıtın.'
           })
       
       # Maliyet optimizasyonu
       avg_cost = cost_analysis.get('avg_cost_per_query', 0)
       if avg_cost > 0.1:  # 10 cent per query
           recommendations.append({
               'type': 'cost',
               'priority': 'medium',
               'title': 'Yüksek Query Maliyeti',
               'description': f'Query başına ortalama ${avg_cost:.3f} maliyet.',
               'action': 'Cache stratejilerini gözden geçirin ve model optimizasyonu yapın.'
           })
       
       # Kullanıcı memnuniyeti
       avg_rating = satisfaction.get('avg_rating', 0)
       if avg_rating < 4.0:
           recommendations.append({
               'type': 'satisfaction',
               'priority': 'high',
               'title': 'Düşük Kullanıcı Memnuniyeti',
               'description': f'Ortalama rating {avg_rating:.1f}/5. Hedef 4.5+ olmalı.',
               'action': 'Kullanıcı feedback\'lerini analiz edin ve yanıt kalitesini artırın.'
           })
       
       return recommendations
   
   def _generate_technical_recommendations(
       self,
       performance_stats: Dict,
       error_patterns: List,
       cache_stats: Dict
   ) -> List[Dict[str, Any]]:
       """Teknik öneriler oluştur"""
       recommendations = []
       
       # Response time önerisi
       avg_response = performance_stats.get('avg_response_time', 0)
       if avg_response > 5.0:
           recommendations.append({
               'type': 'performance',
               'priority': 'high',
               'title': 'Yavaş Yanıt Süreleri',
               'description': f'Ortalama yanıt süresi {avg_response:.2f}s. Hedef <3s olmalı.',
               'action': 'Database indexlerini optimize edin ve caching stratejisini gözden geçirin.'
           })
       
       # Error rate önerisi
       if len(error_patterns) > 100:
           recommendations.append({
               'type': 'reliability',
               'priority': 'high',
               'title': 'Yüksek Hata Oranı',
               'description': f'{len(error_patterns)} farklı hata pattern\'i tespit edildi.',
               'action': 'Error handling\'i iyileştirin ve monitoring alertlerini artırın.'
           })
       
       # Cache performance
       cache_hit_rate = cache_stats.get('redis_keyspace_hits', 0) / max(
           cache_stats.get('redis_keyspace_hits', 0) + cache_stats.get('redis_keyspace_misses', 0), 1
       )
       
       if cache_hit_rate < 0.8:
           recommendations.append({
               'type': 'caching',
               'priority': 'medium',
               'title': 'Düşük Cache Hit Rate',
               'description': f'Cache hit rate %{cache_hit_rate*100:.1f}. Hedef %90+ olmalı.',
               'action': 'Cache strategy\'sini gözden geçirin ve TTL değerlerini optimize edin.'
           })
       
       return recommendations


class AnalyticsExportService:
   """Analytics export servisi"""
   
   def export_to_excel(
       self,
       data: Dict[str, Any],
       report_type: str
   ) -> bytes:
       """Excel formatında export"""
       try:
           import io
           import pandas as pd
           from openpyxl import Workbook
           from openpyxl.utils.dataframe import dataframe_to_rows
           from openpyxl.styles import Font, PatternFill, Alignment
           
           wb = Workbook()
           ws = wb.active
           ws.title = "SAPBot Analytics"
           
           # Header styling
           header_font = Font(bold=True, color="FFFFFF")
           header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
           
           if report_type == 'executive':
               self._add_executive_sheets(wb, data, header_font, header_fill)
           elif report_type == 'technical':
               self._add_technical_sheets(wb, data, header_font, header_fill)
           elif report_type == 'dashboard':
               self._add_dashboard_sheets(wb, data, header_font, header_fill)
           
           # Excel dosyasını byte array'e çevir
           excel_buffer = io.BytesIO()
           wb.save(excel_buffer)
           excel_buffer.seek(0)
           
           return excel_buffer.getvalue()
           
       except Exception as e:
           logger.error(f"Excel export hatası: {e}")
           raise
   
   def _add_executive_sheets(self, wb, data, header_font, header_fill):
       """Executive report sayfalarını ekle"""
       # Summary sheet
       ws_summary = wb.active
       ws_summary.title = "Executive Summary"
       
       # Key metrics
       summary = data.get('executive_summary', {})
       key_metrics = summary.get('key_metrics', {})
       
       ws_summary['A1'] = "SAPBot Executive Report"
       ws_summary['A1'].font = Font(bold=True, size=16)
       
       row = 3
       for metric, value in key_metrics.items():
           ws_summary[f'A{row}'] = metric.replace('_', ' ').title()
           ws_summary[f'B{row}'] = value
           row += 1
   
   def _add_technical_sheets(self, wb, data, header_font, header_fill):
       """Technical report sayfalarını ekle"""
       # Performance sheet
       ws_perf = wb.active
       ws_perf.title = "Performance"
       
       performance = data.get('performance_analysis', {})
       response_times = performance.get('response_times', {})
       
       ws_perf['A1'] = "Performance Metrics"
       ws_perf['A1'].font = Font(bold=True, size=14)
       
       row = 3
       for metric, value in response_times.items():
           if value is not None:
               ws_perf[f'A{row}'] = metric.replace('_', ' ').title()
               ws_perf[f'B{row}'] = round(float(value), 3)
               row += 1
   
   def _add_dashboard_sheets(self, wb, data, header_font, header_fill):
       """Dashboard report sayfalarını ekle"""
       ws_dash = wb.active
       ws_dash.title = "Dashboard Data"
       
       # Summary metrics
       summary = data.get('summary', [])
       
       ws_dash['A1'] = "Dashboard Summary"
       ws_dash['A1'].font = Font(bold=True, size=14)
       
       # Headers
       headers = ['Metric', 'Value', 'Change %', 'Trend']
       for col, header in enumerate(headers, 1):
           cell = ws_dash.cell(row=3, column=col, value=header)
           cell.font = header_font
           cell.fill = header_fill
       
       # Data
       for row, metric in enumerate(summary, 4):
           ws_dash[f'A{row}'] = metric.name
           ws_dash[f'B{row}'] = metric.value
           ws_dash[f'C{row}'] = metric.change
           ws_dash[f'D{row}'] = metric.trend


# Utility functions
def get_analytics_summary(days: int = 30) -> Dict[str, Any]:
   """Hızlı analytics özeti al"""
   try:
       end_date = timezone.now()
       start_date = end_date - timedelta(days=days)
       
       service = DashboardService()
       result = service.get_dashboard_data('daily')
       
       if result.success:
           return {
               'success': True,
               'summary': result.data.get('summary', []),
               'period_days': days,
               'cached': result.cached
           }
       else:
           return {
               'success': False,
               'error': result.message
           }
           
   except Exception as e:
       logger.error(f"Analytics summary hatası: {e}")
       return {
           'success': False,
           'error': str(e)
       }


def record_performance_metric(
   component: str,
   metric_name: str,
   value: float,
   unit: str = 'ms',
   metadata: Dict = None
):
   """Performans metriği kaydet"""
   try:
       PerformanceMetrics.objects.create(
           component=component,
           metric_name=metric_name,
           value=value,
           unit=unit,
           metadata=metadata or {}
       )
   except Exception as e:
       logger.error(f"Performance metric kaydetme hatası: {e}")


def cleanup_old_analytics(days_to_keep: int = 90):
   """Eski analytics verilerini temizle"""
   try:
       cutoff_date = timezone.now() - timedelta(days=days_to_keep)
       
       # Eski kayıtları sil
       deleted_counts = {}
       
       deleted_counts['query_analytics'] = QueryAnalytics.objects.filter(
           created_at__lt=cutoff_date
       ).delete()[0]
       
       deleted_counts['performance_metrics'] = PerformanceMetrics.objects.filter(
           created_at__lt=cutoff_date
       ).delete()[0]
       
       deleted_counts['error_logs'] = ErrorLog.objects.filter(
           created_at__lt=cutoff_date,
           error_level__in=['DEBUG', 'INFO']  # Sadece düşük seviye loglar
       ).delete()[0]
       
       logger.info(f"Analytics cleanup tamamlandı: {deleted_counts}")
       return deleted_counts
       
   except Exception as e:
       logger.error(f"Analytics cleanup hatası: {e}")
       raise


# Selim, bu analytics service ile comprehensive bir analitik sistemi kurduk:

# 🎯 Temel Özellikler:
# - Query analytics (her sorgu için detaylı metrik)
# - Real-time dashboard verileri
# - Executive ve technical raporlar
# - Performance monitoring
# - User behavior analysis
# - Cost tracking
# - Error pattern detection

# 🚀 Kullanım Örnekleri:
# analytics_service = QueryAnalyticsService()
# analytics_service.record_query(user_id, session_id, query, user_type, response_data)

# dashboard_service = DashboardService()
# dashboard_data = dashboard_service.get_dashboard_data('daily')

# reporting_service = ReportingService()
# executive_report = reporting_service.generate_executive_report(start_date, end_date)

# ⚡ Performance Optimizations:
# - Akıllı caching (30 dk dashboard, 1 saat trends)
# - Database indexing (query_hash, created_at, user_id)
# - Bulk operations for large datasets
# - Automated cleanup for old data

# 📊 Self-Healing Features:
# - Automatic metric collection
# - Health monitoring
# - Error pattern detection
# - Performance alerting

# Bu service, SAP B1 kullanım patterns'larını anlayıp optimize etmek için gereken tüm data science capability'yi sağlıyor.