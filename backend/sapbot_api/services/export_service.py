# backend/sapbot_api/services/export_service.py
"""
SAPBot API Export Service

Bu servis chat geçmişleri, dökümanlar, analitik veriler ve 
sistem loglarının farklı formatlarda export edilmesini sağlar.
"""

import os
import io
import csv
import json
import zipfile
from typing import Dict, List, Any, Optional, Union, BinaryIO
from datetime import datetime, timedelta
from decimal import Decimal
from dataclasses import dataclass, asdict
import logging

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils import timezone
from django.db.models import QuerySet, Q, Count, Avg, Sum
from django.http import HttpResponse, StreamingHttpResponse

from ..models import (
   ChatConversation, ChatMessage, DocumentSource, KnowledgeChunk,
   QueryAnalytics, UsageStatistics, UserFeedback, ErrorLog,
   UserProfile, SystemLog
)
from ..utils.exceptions import ExportException, ValidationException
from ..utils.helpers import format_file_size, time_ago, format_currency
from ..utils.cache_utils import CacheManager
from ..utils.validators import UserInputValidator

logger = logging.getLogger(__name__)


@dataclass
class ExportRequest:
   """Export request modeli"""
   export_type: str
   format: str
   start_date: Optional[datetime] = None
   end_date: Optional[datetime] = None
   filters: Dict[str, Any] = None
   user_id: Optional[str] = None
   include_metadata: bool = True
   compress: bool = False
   
   def __post_init__(self):
       if self.filters is None:
           self.filters = {}


@dataclass
class ExportResult:
   """Export sonuç modeli"""
   file_path: str
   file_size: int
   record_count: int
   export_type: str
   format: str
   created_at: datetime
   expires_at: datetime
   download_url: str
   metadata: Dict[str, Any] = None


class BaseExporter:
   """Temel export sınıfı"""
   
   def __init__(self):
       self.cache = CacheManager()
       self.temp_dir = getattr(settings, 'EXPORT_TEMP_DIR', '/tmp/exports')
       os.makedirs(self.temp_dir, exist_ok=True)
   
   def validate_request(self, request: ExportRequest) -> ExportRequest:
       """Export request doğrulama"""
       if not request.export_type:
           raise ValidationException("Export type gerekli", field="export_type")
       
       if not request.format:
           raise ValidationException("Export format gerekli", field="format")
       
       # Tarih validasyonu
       if request.start_date and request.end_date:
           if request.start_date >= request.end_date:
               raise ValidationException("Başlangıç tarihi bitiş tarihinden önce olmalı")
           
           # Maksimum 1 yıl
           if (request.end_date - request.start_date).days > 365:
               raise ValidationException("Export aralığı maksimum 1 yıl olabilir")
       
       return request
   
   def get_filename(self, export_type: str, format: str, timestamp: str = None) -> str:
       """Dosya adı oluştur"""
       if not timestamp:
           timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
       
       return f"sapbot_{export_type}_{timestamp}.{format.lower()}"
   
   def save_to_storage(self, file_content: bytes, filename: str) -> str:
       """Dosyayı storage'a kaydet"""
       file_path = f"exports/{datetime.now().strftime('%Y/%m/%d')}/{filename}"
       
       content_file = ContentFile(file_content)
       saved_path = default_storage.save(file_path, content_file)
       
       return saved_path


class ChatExporter(BaseExporter):
   """Chat export sınıfı"""
   
   def export_conversations(self, request: ExportRequest) -> ExportResult:
       """Konuşmaları export et"""
       try:
           request = self.validate_request(request)
           
           # Veri sorgulama
           conversations = self._get_conversations(request)
           
           if request.format.lower() == 'csv':
               return self._export_to_csv(conversations, request)
           elif request.format.lower() == 'excel':
               return self._export_to_excel(conversations, request)
           elif request.format.lower() == 'json':
               return self._export_to_json(conversations, request)
           elif request.format.lower() == 'pdf':
               return self._export_to_pdf(conversations, request)
           else:
               raise ValidationException(f"Desteklenmeyen format: {request.format}")
               
       except Exception as e:
           logger.error(f"Chat export hatası: {e}")
           raise ExportException(f"Chat export başarısız: {str(e)}")
   
   def _get_conversations(self, request: ExportRequest) -> QuerySet:
       """Konuşmaları al"""
       queryset = ChatConversation.objects.select_related(
           'user'
       ).prefetch_related(
           'messages__sources_used'
       )
       
       # Tarih filtresi
       if request.start_date:
           queryset = queryset.filter(created_at__gte=request.start_date)
       if request.end_date:
           queryset = queryset.filter(created_at__lte=request.end_date)
       
       # Kullanıcı filtresi
       if request.user_id:
           queryset = queryset.filter(user_id=request.user_id)
       
       # Diğer filtreler
       filters = request.filters
       if filters.get('user_type'):
           queryset = queryset.filter(user_type=filters['user_type'])
       
       if filters.get('has_messages'):
           queryset = queryset.annotate(
               message_count=Count('messages')
           ).filter(message_count__gt=0)
       
       return queryset.order_by('-created_at')
   
   def _export_to_csv(self, conversations: QuerySet, request: ExportRequest) -> ExportResult:
       """CSV export"""
       output = io.StringIO()
       writer = csv.writer(output)
       
       # Headers
       headers = [
           'Konuşma ID', 'Kullanıcı Email', 'Kullanıcı Tipi', 'Oturum ID',
           'Mesaj Sayısı', 'Başlangıç Zamanı', 'Son Aktivite', 'Süre (dk)',
           'IP Adresi', 'Tarayıcı'
       ]
       
       if request.include_metadata:
           headers.extend(['Metadata', 'Arşivlenmiş'])
       
       writer.writerow(headers)
       
       # Data
       record_count = 0
       for conv in conversations:
           row = [
               str(conv.id),
               conv.user.email if conv.user else 'Anonim',
               conv.get_user_type_display(),
               conv.session_id,
               conv.message_count,
               conv.created_at.strftime('%Y-%m-%d %H:%M:%S'),
               conv.last_activity.strftime('%Y-%m-%d %H:%M:%S'),
               conv.duration_minutes,
               conv.ip_address or '',
               conv.user_agent or ''
           ]
           
           if request.include_metadata:
               row.extend([
                   json.dumps(conv.metadata, ensure_ascii=False),
                   'Evet' if conv.is_archived else 'Hayır'
               ])
           
           writer.writerow(row)
           record_count += 1
       
       # Dosya kaydet
       content = output.getvalue().encode('utf-8-sig')  # BOM ile UTF-8
       filename = self.get_filename('conversations', 'csv')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=record_count,
           export_type='conversations',
           format='csv',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_to_excel(self, conversations: QuerySet, request: ExportRequest) -> ExportResult:
       """Excel export"""
       # DataFrame oluştur
       data = []
       for conv in conversations:
           row = {
               'Konuşma ID': str(conv.id),
               'Kullanıcı Email': conv.user.email if conv.user else 'Anonim',
               'Kullanıcı Tipi': conv.get_user_type_display(),
               'Oturum ID': conv.session_id,
               'Mesaj Sayısı': conv.message_count,
               'Başlangıç Zamanı': conv.created_at,
               'Son Aktivite': conv.last_activity,
               'Süre (dk)': conv.duration_minutes,
               'IP Adresi': conv.ip_address or '',
               'Tarayıcı': conv.user_agent or ''
           }
           
           if request.include_metadata:
               row.update({
                   'Metadata': json.dumps(conv.metadata, ensure_ascii=False),
                   'Arşivlenmiş': 'Evet' if conv.is_archived else 'Hayır'
               })
           
           data.append(row)
       
       df = pd.DataFrame(data)
       
       # Excel dosyası oluştur
       output = io.BytesIO()
       with pd.ExcelWriter(output, engine='openpyxl') as writer:
           df.to_excel(writer, sheet_name='Konuşmalar', index=False)
           
           # Formatting
           worksheet = writer.sheets['Konuşmalar']
           
           # Header style
           header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
           header_font = Font(color='FFFFFF', bold=True)
           
           for cell in worksheet[1]:
               cell.fill = header_fill
               cell.font = header_font
               cell.alignment = Alignment(horizontal='center')
           
           # Auto-width
           for column in worksheet.columns:
               max_length = 0
               column_letter = column[0].column_letter
               for cell in column:
                   if len(str(cell.value)) > max_length:
                       max_length = len(str(cell.value))
               adjusted_width = min(max_length + 2, 50)
               worksheet.column_dimensions[column_letter].width = adjusted_width
       
       content = output.getvalue()
       filename = self.get_filename('conversations', 'xlsx')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=len(data),
           export_type='conversations',
           format='excel',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_to_json(self, conversations: QuerySet, request: ExportRequest) -> ExportResult:
       """JSON export"""
       data = {
           'export_info': {
               'type': 'conversations',
               'format': 'json',
               'created_at': timezone.now().isoformat(),
               'filters': request.filters,
               'include_metadata': request.include_metadata
           },
           'conversations': []
       }
       
       for conv in conversations:
           conv_data = {
               'id': str(conv.id),
               'user': {
                   'email': conv.user.email if conv.user else None,
                   'type': conv.user_type
               },
               'session_id': conv.session_id,
               'message_count': conv.message_count,
               'duration_minutes': conv.duration_minutes,
               'created_at': conv.created_at.isoformat(),
               'last_activity': conv.last_activity.isoformat(),
               'ip_address': conv.ip_address,
               'user_agent': conv.user_agent,
               'is_archived': conv.is_archived
           }
           
           if request.include_metadata:
               conv_data['metadata'] = conv.metadata
           
           # Mesajları da dahil et
           if request.filters.get('include_messages'):
               conv_data['messages'] = []
               for msg in conv.messages.all():
                   msg_data = {
                       'id': str(msg.id),
                       'type': msg.message_type,
                       'content': msg.content,
                       'created_at': msg.created_at.isoformat(),
                       'response_time': msg.response_time,
                       'token_count': msg.token_count,
                       'sources_count': msg.source_count
                   }
                   conv_data['messages'].append(msg_data)
           
           data['conversations'].append(conv_data)
       
       # JSON string'e çevir
       content = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
       filename = self.get_filename('conversations', 'json')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=len(data['conversations']),
           export_type='conversations',
           format='json',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_to_pdf(self, conversations: QuerySet, request: ExportRequest) -> ExportResult:
       """PDF export"""
       output = io.BytesIO()
       doc = SimpleDocTemplate(output, pagesize=A4)
       
       # Styles
       styles = getSampleStyleSheet()
       title_style = ParagraphStyle(
           'CustomTitle',
           parent=styles['Title'],
           fontSize=16,
           textColor=colors.navy,
           spaceAfter=30
       )
       
       story = []
       
       # Başlık
       title = Paragraph("SAPBot Chat Konuşmaları Raporu", title_style)
       story.append(title)
       
       # Özet bilgiler
       summary_data = [
           ['Toplam Konuşma', str(conversations.count())],
           ['Export Tarihi', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
           ['Tarih Aralığı', f"{request.start_date or 'Başlangıç'} - {request.end_date or 'Son'}"]
       ]
       
       summary_table = Table(summary_data)
       summary_table.setStyle(TableStyle([
           ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
           ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
           ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
           ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
           ('FONTSIZE', (0, 0), (-1, -1), 10),
           ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
       ]))
       
       story.append(summary_table)
       story.append(Spacer(1, 20))
       
       # Konuşma listesi
       table_data = [['Kullanıcı', 'Oturum ID', 'Mesaj Sayısı', 'Tarih', 'Süre']]
       
       for conv in conversations[:100]:  # İlk 100 konuşma
           table_data.append([
               conv.user.email if conv.user else 'Anonim',
               conv.session_id[:20] + '...' if len(conv.session_id) > 20 else conv.session_id,
               str(conv.message_count),
               conv.created_at.strftime('%Y-%m-%d'),
               f"{conv.duration_minutes:.0f} dk"
           ])
       
       table = Table(table_data)
       table.setStyle(TableStyle([
           ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
           ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
           ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
           ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
           ('FONTSIZE', (0, 0), (-1, 0), 10),
           ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
           ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
           ('GRID', (0, 0), (-1, -1), 1, colors.black)
       ]))
       
       story.append(table)
       
       # PDF oluştur
       doc.build(story)
       content = output.getvalue()
       
       filename = self.get_filename('conversations', 'pdf')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=conversations.count(),
           export_type='conversations',
           format='pdf',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )


class DocumentExporter(BaseExporter):
   """Döküman export sınıfı"""
   
   def export_documents(self, request: ExportRequest) -> ExportResult:
       """Dökümanları export et"""
       try:
           request = self.validate_request(request)
           
           documents = self._get_documents(request)
           
           if request.format.lower() == 'csv':
               return self._export_documents_csv(documents, request)
           elif request.format.lower() == 'excel':
               return self._export_documents_excel(documents, request)
           elif request.format.lower() == 'zip':
               return self._export_documents_zip(documents, request)
           else:
               raise ValidationException(f"Desteklenmeyen format: {request.format}")
               
       except Exception as e:
           logger.error(f"Document export hatası: {e}")
           raise ExportException(f"Document export başarısız: {str(e)}")
   
   def _get_documents(self, request: ExportRequest) -> QuerySet:
       """Dökümanları al"""
       queryset = DocumentSource.objects.select_related(
           'uploaded_by'
       ).prefetch_related(
           'chunks'
       )
       
       # Tarih filtresi
       if request.start_date:
           queryset = queryset.filter(created_at__gte=request.start_date)
       if request.end_date:
           queryset = queryset.filter(created_at__lte=request.end_date)
       
       # Filtreler
       filters = request.filters
       if filters.get('document_type'):
           queryset = queryset.filter(document_type=filters['document_type'])
       
       if filters.get('language'):
           queryset = queryset.filter(language=filters['language'])
       
       if filters.get('processing_status'):
           queryset = queryset.filter(processing_status=filters['processing_status'])
       
       if filters.get('only_active'):
           queryset = queryset.filter(is_active=True)
       
       return queryset.order_by('-created_at')
   
   def _export_documents_csv(self, documents: QuerySet, request: ExportRequest) -> ExportResult:
       """Dökümanları CSV'ye export et"""
       output = io.StringIO()
       writer = csv.writer(output)
       
       # Headers
       headers = [
           'ID', 'Başlık', 'Tip', 'Dil', 'Durum', 'Boyut (MB)',
           'Chunk Sayısı', 'Yükleyen', 'Yükleme Tarihi', 'İşlenme Tarihi'
       ]
       
       if request.include_metadata:
           headers.extend(['Etiketler', 'Metadata', 'Dosya Hash'])
       
       writer.writerow(headers)
       
       # Data
       record_count = 0
       for doc in documents:
           row = [
               str(doc.id),
               doc.title,
               doc.get_document_type_display(),
               doc.get_language_display(),
               doc.get_processing_status_display(),
               doc.file_size_mb,
               doc.chunk_count,
               doc.uploaded_by.email if doc.uploaded_by else '',
               doc.created_at.strftime('%Y-%m-%d %H:%M:%S'),
               doc.processed_at.strftime('%Y-%m-%d %H:%M:%S') if doc.processed_at else ''
           ]
           
           if request.include_metadata:
               row.extend([
                   ', '.join(doc.tags),
                   json.dumps(doc.metadata, ensure_ascii=False),
                   doc.file_hash or ''
               ])
           
           writer.writerow(row)
           record_count += 1
       
       content = output.getvalue().encode('utf-8-sig')
       filename = self.get_filename('documents', 'csv')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=record_count,
           export_type='documents',
           format='csv',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_documents_excel(self, documents: QuerySet, request: ExportRequest) -> ExportResult:
       """Dökümanları Excel'e export et"""
       # Multiple sheets
       output = io.BytesIO()
       
       with pd.ExcelWriter(output, engine='openpyxl') as writer:
           # Ana döküman bilgileri
           doc_data = []
           for doc in documents:
               row = {
                   'ID': str(doc.id),
                   'Başlık': doc.title,
                   'Tip': doc.get_document_type_display(),
                   'Dil': doc.get_language_display(),
                   'Durum': doc.get_processing_status_display(),
                   'Boyut (MB)': doc.file_size_mb,
                   'Chunk Sayısı': doc.chunk_count,
                   'Yükleyen': doc.uploaded_by.email if doc.uploaded_by else '',
                   'Yükleme Tarihi': doc.created_at,
                   'İşlenme Tarihi': doc.processed_at
               }
               doc_data.append(row)
           
           df_docs = pd.DataFrame(doc_data)
           df_docs.to_excel(writer, sheet_name='Dökümanlar', index=False)
           
           # İstatistikler sheet'i
           stats_data = [
               ['Toplam Döküman', documents.count()],
               ['Tamamlanmış', documents.filter(processing_status='completed').count()],
               ['İşleniyor', documents.filter(processing_status='processing').count()],
               ['Başarısız', documents.filter(processing_status='failed').count()],
               ['Toplam Boyut (MB)', sum(doc.file_size_mb for doc in documents)],
               ['Toplam Chunk', sum(doc.chunk_count for doc in documents)]
           ]
           
           df_stats = pd.DataFrame(stats_data, columns=['Metrik', 'Değer'])
           df_stats.to_excel(writer, sheet_name='İstatistikler', index=False)
           
           # Formatting
           for sheet_name in writer.sheets:
               worksheet = writer.sheets[sheet_name]
               
               # Header formatting
               header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
               header_font = Font(color='FFFFFF', bold=True)
               
               for cell in worksheet[1]:
                   cell.fill = header_fill
                   cell.font = header_font
                   cell.alignment = Alignment(horizontal='center')
               
               # Auto-width
               for column in worksheet.columns:
                   max_length = 0
                   column_letter = column[0].column_letter
                   for cell in column:
                       if len(str(cell.value)) > max_length:
                           max_length = len(str(cell.value))
                   adjusted_width = min(max_length + 2, 50)
                   worksheet.column_dimensions[column_letter].width = adjusted_width
       
       content = output.getvalue()
       filename = self.get_filename('documents', 'xlsx')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=len(doc_data),
           export_type='documents',
           format='excel',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_documents_zip(self, documents: QuerySet, request: ExportRequest) -> ExportResult:
       """Dökümanları ZIP olarak export et"""
       output = io.BytesIO()
       
       with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zipf:
           # Manifest dosyası
           manifest = {
               'export_info': {
                   'type': 'documents_with_files',
                   'created_at': timezone.now().isoformat(),
                   'total_documents': documents.count()
               },
               'documents': []
           }
           
           for doc in documents:
               doc_info = {
                   'id': str(doc.id),
                   'title': doc.title,
                   'filename': os.path.basename(doc.file_path.name) if doc.file_path else None,
                   'type': doc.document_type,
                   'language': doc.language,
                   'status': doc.processing_status,
                   'size_mb': doc.file_size_mb,
                   'chunk_count': doc.chunk_count,
                   'created_at': doc.created_at.isoformat()
               }
               
               manifest['documents'].append(doc_info)
               
               # Dosyayı ZIP'e ekle
               if doc.file_path and default_storage.exists(doc.file_path.name):
                   try:
                       file_content = doc.file_path.read()
                       safe_filename = f"{doc.id}_{os.path.basename(doc.file_path.name)}"
                       zipf.writestr(f"files/{safe_filename}", file_content)
                   except Exception as e:
                       logger.warning(f"Dosya ZIP'e eklenemedi: {doc.file_path.name} - {e}")
           
           # Manifest'i ekle
           manifest_json = json.dumps(manifest, ensure_ascii=False, indent=2)
           zipf.writestr('manifest.json', manifest_json)
           
           # README dosyası
           readme_content = """SAPBot Döküman Export Paketi

Bu ZIP dosyası şunları içerir:
- manifest.json: Tüm döküman bilgileri
- files/ klasörü: Orijinal döküman dosyaları

Her dosya şu format ile adlandırılmıştır:
{document_id}_{original_filename}

Manifest dosyasındaki bilgilerle dosyaları eşleştirebilirsiniz.
"""
           zipf.writestr('README.txt', readme_content)
       
       content = output.getvalue()
       filename = self.get_filename('documents_full', 'zip')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=documents.count(),
           export_type='documents',
           format='zip',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path),
           metadata={'includes_files': True}
       )


class AnalyticsExporter(BaseExporter):
   """Analitik export sınıfı"""
   
   def export_analytics(self, request: ExportRequest) -> ExportResult:
       """Analitik verileri export et"""
       try:
           request = self.validate_request(request)
           
           if request.export_type == 'query_analytics':
               return self._export_query_analytics(request)
           elif request.export_type == 'usage_statistics':
               return self._export_usage_statistics(request)
           elif request.export_type == 'user_feedback':
               return self._export_user_feedback(request)
           elif request.export_type == 'comprehensive':
               return self._export_comprehensive_analytics(request)
           else:
               raise ValidationException(f"Desteklenmeyen analitik tipi: {request.export_type}")
               
       except Exception as e:
           logger.error(f"Analytics export hatası: {e}")
           raise ExportException(f"Analytics export başarısız: {str(e)}")
   
   def _export_query_analytics(self, request: ExportRequest) -> ExportResult:
       """Sorgu analitiklerini export et"""
       queryset = QueryAnalytics.objects.select_related('user')
       
       # Tarih filtresi
       if request.start_date:
           queryset = queryset.filter(created_at__gte=request.start_date)
       if request.end_date:
           queryset = queryset.filter(created_at__lte=request.end_date)
       
       # Filtreler
       filters = request.filters
       if filters.get('sap_module'):
           queryset = queryset.filter(sap_module_detected=filters['sap_module'])
       
       if filters.get('user_type'):
           queryset = queryset.filter(user_type=filters['user_type'])
       
       if filters.get('intent'):
           queryset = queryset.filter(intent_detected=filters['intent'])
       
       queryset = queryset.order_by('-created_at')
       
       if request.format.lower() == 'excel':
           return self._export_query_analytics_excel(queryset, request)
       else:
           return self._export_query_analytics_csv(queryset, request)
   
   def _export_query_analytics_excel(self, queryset: QuerySet, request: ExportRequest) -> ExportResult:
       """Query analytics Excel export"""
       output = io.BytesIO()
       
       with pd.ExcelWriter(output, engine='openpyxl') as writer:
           # Ana veri
           analytics_data = []
           for qa in queryset:
               row = {
                   'Tarih': qa.created_at,
                   'Kullanıcı': qa.user.email if qa.user else 'Anonim',
                   'Kullanıcı Tipi': qa.get_user_type_display(),
                   'Sorgu': qa.query[:100] + '...' if len(qa.query) > 100 else qa.query,
                   'Sorgu Uzunluğu': qa.query_length,
                   'SAP Modülü': qa.sap_module_detected or '',
                   'Niyet': qa.get_intent_detected_display() if qa.intent_detected else '',
                   'Güven Skoru': qa.confidence_score,
                   'Yanıt Oluşturuldu': 'Evet' if qa.response_generated else 'Hayır',
                   'Yanıt Süresi (sn)': qa.response_time,
                   'Kaynak Sayısı': qa.sources_used_count,
                   'Token Sayısı': qa.tokens_used,
                   'Maliyet (₺)': qa.cost_estimate,
                   'Hata': 'Evet' if qa.error_occurred else 'Hayır',
                   'IP Adresi': qa.ip_address or '',
                   'Dil': qa.language_detected or ''
               }
               analytics_data.append(row)
           
           df_analytics = pd.DataFrame(analytics_data)
           df_analytics.to_excel(writer, sheet_name='Sorgu Analitikleri', index=False)
           
           # Özet istatistikler
           summary_data = self._calculate_analytics_summary(queryset)
           df_summary = pd.DataFrame(list(summary_data.items()), columns=['Metrik', 'Değer'])
           df_summary.to_excel(writer, sheet_name='Özet İstatistikler', index=False)
           
           # SAP Modül analizi
           module_stats = queryset.values('sap_module_detected').annotate(
               count=Count('id'),
               avg_response_time=Avg('response_time'),
               success_rate=Avg('response_generated')
           ).order_by('-count')
           
           module_data = []
           for stat in module_stats:
               module_data.append({
                   'SAP Modülü': stat['sap_module_detected'] or 'Belirtilmemiş',
                   'Sorgu Sayısı': stat['count'],
                   'Ort. Yanıt Süresi': round(stat['avg_response_time'] or 0, 2),
                   'Başarı Oranı (%)': round((stat['success_rate'] or 0) * 100, 1)
               })
           
           df_modules = pd.DataFrame(module_data)
           df_modules.to_excel(writer, sheet_name='SAP Modül Analizi', index=False)
           
           # Intent analizi
           intent_stats = queryset.values('intent_detected').annotate(
               count=Count('id'),
               avg_confidence=Avg('confidence_score')
           ).order_by('-count')
           
           intent_data = []
           for stat in intent_stats:
               intent_data.append({
                   'Niyet': stat['intent_detected'] or 'Belirtilmemiş',
                   'Sorgu Sayısı': stat['count'],
                   'Ort. Güven Skoru': round(stat['avg_confidence'] or 0, 3)
               })
           
           df_intents = pd.DataFrame(intent_data)
           df_intents.to_excel(writer, sheet_name='Niyet Analizi', index=False)
           
           # Formatting tüm sheet'ler için
           self._format_excel_sheets(writer)
       
       content = output.getvalue()
       filename = self.get_filename('query_analytics', 'xlsx')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=queryset.count(),
           export_type='query_analytics',
           format='excel',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_usage_statistics(self, request: ExportRequest) -> ExportResult:
       """Kullanım istatistiklerini export et"""
       queryset = UsageStatistics.objects.all()
       
       # Tarih filtresi
       if request.start_date:
           queryset = queryset.filter(date__gte=request.start_date.date())
       if request.end_date:
           queryset = queryset.filter(date__lte=request.end_date.date())
       
       # Metrik tipi filtresi
       if request.filters.get('metric_type'):
           queryset = queryset.filter(metric_type=request.filters['metric_type'])
       
       queryset = queryset.order_by('-date')
       
       output = io.BytesIO()
       
       with pd.ExcelWriter(output, engine='openpyxl') as writer:
           # Ana kullanım verileri
           usage_data = []
           for stat in queryset:
               row = {
                   'Tarih': stat.date,
                   'Metrik Tipi': stat.get_metric_type_display(),
                   'Toplam Sorgu': stat.total_queries,
                   'Başarılı Sorgu': stat.successful_queries,
                   'Başarısız Sorgu': stat.failed_queries,
                   'Başarı Oranı (%)': stat.success_rate,
                   'Benzersiz Kullanıcı': stat.unique_users,
                   'Benzersiz Oturum': stat.unique_sessions,
                   'Ort. Yanıt Süresi (sn)': stat.avg_response_time,
                   'Ort. Memnuniyet': stat.avg_satisfaction,
                   'Kullanılan Token': stat.total_tokens_used,
                   'Toplam Maliyet (₺)': stat.total_cost,
                   'İşlenen Döküman': stat.documents_processed,
                   'Oluşturulan Chunk': stat.chunks_created,
                   'Kullanıcı Başına Ort. Sorgu': stat.avg_queries_per_user
               }
               usage_data.append(row)
           
           df_usage = pd.DataFrame(usage_data)
           df_usage.to_excel(writer, sheet_name='Kullanım İstatistikleri', index=False)
           
           # Trend analizi (günlük veriler için)
           daily_stats = queryset.filter(metric_type='daily').order_by('date')
           if daily_stats.exists():
               trend_data = []
               for i, stat in enumerate(daily_stats):
                   prev_stat = daily_stats[i-1] if i > 0 else None
                   
                   query_growth = 0
                   user_growth = 0
                   
                   if prev_stat:
                       if prev_stat.total_queries > 0:
                           query_growth = ((stat.total_queries - prev_stat.total_queries) / prev_stat.total_queries) * 100
                       if prev_stat.unique_users > 0:
                           user_growth = ((stat.unique_users - prev_stat.unique_users) / prev_stat.unique_users) * 100
                   
                   row = {
                       'Tarih': stat.date,
                       'Sorgu Sayısı': stat.total_queries,
                       'Sorgu Büyüme (%)': round(query_growth, 2),
                       'Kullanıcı Sayısı': stat.unique_users,
                       'Kullanıcı Büyüme (%)': round(user_growth, 2),
                       'Başarı Oranı (%)': round(stat.success_rate, 2),
                       'Ort. Yanıt Süresi': round(stat.avg_response_time, 2)
                   }
                   trend_data.append(row)
               
               df_trends = pd.DataFrame(trend_data)
               df_trends.to_excel(writer, sheet_name='Trend Analizi', index=False)
           
           # Top SAP modülleri
           all_modules = []
           for stat in queryset:
               if stat.top_sap_modules:
                   for module_data in stat.top_sap_modules:
                       all_modules.append({
                           'Tarih': stat.date,
                           'SAP Modülü': module_data.get('module', ''),
                           'Kullanım Sayısı': module_data.get('count', 0)
                       })
           
           if all_modules:
               df_modules = pd.DataFrame(all_modules)
               module_summary = df_modules.groupby('SAP Modülü').agg({
                   'Kullanım Sayısı': 'sum'
               }).reset_index().sort_values('Kullanım Sayısı', ascending=False)
               
               module_summary.to_excel(writer, sheet_name='En Çok Kullanılan SAP Modülleri', index=False)
           
           self._format_excel_sheets(writer)
       
       content = output.getvalue()
       filename = self.get_filename('usage_statistics', 'xlsx')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=queryset.count(),
           export_type='usage_statistics',
           format='excel',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_user_feedback(self, request: ExportRequest) -> ExportResult:
       """Kullanıcı geri bildirimlerini export et"""
       queryset = UserFeedback.objects.select_related('user', 'message')
       
       # Tarih filtresi
       if request.start_date:
           queryset = queryset.filter(created_at__gte=request.start_date)
       if request.end_date:
           queryset = queryset.filter(created_at__lte=request.end_date)
       
       # Filtreler
       filters = request.filters
       if filters.get('feedback_type'):
           queryset = queryset.filter(feedback_type=filters['feedback_type'])
       
       if filters.get('rating'):
           queryset = queryset.filter(rating=filters['rating'])
       
       if filters.get('is_helpful') is not None:
           queryset = queryset.filter(is_helpful=filters['is_helpful'])
       
       queryset = queryset.order_by('-created_at')
       
       output = io.BytesIO()
       
       with pd.ExcelWriter(output, engine='openpyxl') as writer:
           # Ana feedback verileri
           feedback_data = []
           for feedback in queryset:
               row = {
                   'Tarih': feedback.created_at,
                   'Kullanıcı': feedback.user.email if feedback.user else 'Anonim',
                   'Feedback Tipi': feedback.get_feedback_type_display(),
                   'Puan': feedback.rating,
                   'Memnuniyet': feedback.get_satisfaction_display() if feedback.satisfaction else '',
                   'Faydalı mı?': 'Evet' if feedback.is_helpful else ('Hayır' if feedback.is_helpful is False else ''),
                   'Yorum': feedback.comment or '',
                   'İyileştirme Önerileri': feedback.improvement_suggestions or '',
                   'Mesaj ID': str(feedback.message.id) if feedback.message else '',
                   'Oturum ID': feedback.session_id or '',
                   'İşlendi mi?': 'Evet' if feedback.is_processed else 'Hayır',
                   'Yanıt Gönderildi mi?': 'Evet' if feedback.response_sent else 'Hayır'
               }
               feedback_data.append(row)
           
           df_feedback = pd.DataFrame(feedback_data)
           df_feedback.to_excel(writer, sheet_name='Geri Bildirimler', index=False)
           
           # Feedback istatistikleri
           stats_data = [
               ['Toplam Feedback', queryset.count()],
               ['Ortalama Puan', round(queryset.aggregate(Avg('rating'))['rating__avg'] or 0, 2)],
               ['5 Yıldız', queryset.filter(rating=5).count()],
               ['4 Yıldız', queryset.filter(rating=4).count()],
               ['3 Yıldız', queryset.filter(rating=3).count()],
               ['2 Yıldız', queryset.filter(rating=2).count()],
               ['1 Yıldız', queryset.filter(rating=1).count()],
               ['Faydalı Bulan', queryset.filter(is_helpful=True).count()],
               ['Faydalı Bulmayan', queryset.filter(is_helpful=False).count()],
               ['Yorumlu Feedback', queryset.exclude(comment='').exclude(comment__isnull=True).count()],
               ['İşlenmiş Feedback', queryset.filter(is_processed=True).count()],
               ['Yanıtlanmış Feedback', queryset.filter(response_sent=True).count()]
           ]
           
           df_stats = pd.DataFrame(stats_data, columns=['Metrik', 'Değer'])
           df_stats.to_excel(writer, sheet_name='Feedback İstatistikleri', index=False)
           
           # Puan dağılımı zaman içinde (günlük)
           if request.start_date and request.end_date:
               daily_ratings = queryset.extra(
                   select={'day': 'date(created_at)'}
               ).values('day').annotate(
                   avg_rating=Avg('rating'),
                   count=Count('id')
               ).order_by('day')
               
               rating_trend = []
               for item in daily_ratings:
                   rating_trend.append({
                       'Tarih': item['day'],
                       'Ortalama Puan': round(item['avg_rating'], 2),
                       'Feedback Sayısı': item['count']
                   })
               
               if rating_trend:
                   df_trend = pd.DataFrame(rating_trend)
                   df_trend.to_excel(writer, sheet_name='Puan Trendi', index=False)
           
           self._format_excel_sheets(writer)
       
       content = output.getvalue()
       filename = self.get_filename('user_feedback', 'xlsx')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=queryset.count(),
           export_type='user_feedback',
           format='excel',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path)
       )
   
   def _export_comprehensive_analytics(self, request: ExportRequest) -> ExportResult:
       """Kapsamlı analitik raporu"""
       output = io.BytesIO()
       
       with pd.ExcelWriter(output, engine='openpyxl') as writer:
           # Genel özet
           summary_data = self._generate_comprehensive_summary(request)
           df_summary = pd.DataFrame(list(summary_data.items()), columns=['Metrik', 'Değer'])
           df_summary.to_excel(writer, sheet_name='Genel Özet', index=False)
           
           # Kullanıcı performansı
           user_performance = self._get_user_performance_data(request)
           if user_performance:
               df_users = pd.DataFrame(user_performance)
               df_users.to_excel(writer, sheet_name='Kullanıcı Performansı', index=False)
           
           # SAP Modül kullanımı
           module_usage = self._get_module_usage_data(request)
           if module_usage:
               df_modules = pd.DataFrame(module_usage)
               df_modules.to_excel(writer, sheet_name='SAP Modül Kullanımı', index=False)
           
           # Hata analizi
           error_analysis = self._get_error_analysis_data(request)
           if error_analysis:
               df_errors = pd.DataFrame(error_analysis)
               df_errors.to_excel(writer, sheet_name='Hata Analizi', index=False)
           
           # Sistem performansı
           system_performance = self._get_system_performance_data(request)
           if system_performance:
               df_system = pd.DataFrame(system_performance)
               df_system.to_excel(writer, sheet_name='Sistem Performansı', index=False)
           
           self._format_excel_sheets(writer)
       
       content = output.getvalue()
       filename = self.get_filename('comprehensive_analytics', 'xlsx')
       file_path = self.save_to_storage(content, filename)
       
       return ExportResult(
           file_path=file_path,
           file_size=len(content),
           record_count=1,  # Comprehensive report
           export_type='comprehensive',
           format='excel',
           created_at=timezone.now(),
           expires_at=timezone.now() + timedelta(days=7),
           download_url=default_storage.url(file_path),
           metadata={'report_type': 'comprehensive_analytics'}
       )
   
   def _calculate_analytics_summary(self, queryset: QuerySet) -> Dict[str, Any]:
       """Analitik özet hesapla"""
       return {
           'Toplam Sorgu': queryset.count(),
           'Başarılı Yanıt': queryset.filter(response_generated=True).count(),
           'Ortalama Yanıt Süresi (sn)': round(queryset.aggregate(Avg('response_time'))['response_time__avg'] or 0, 2),
           'Ortalama Güven Skoru': round(queryset.aggregate(Avg('confidence_score'))['confidence_score__avg'] or 0, 3),
           'Toplam Token Kullanımı': queryset.aggregate(Sum('tokens_used'))['tokens_used__sum'] or 0,
           'Toplam Maliyet (₺)': queryset.aggregate(Sum('cost_estimate'))['cost_estimate__sum'] or 0,
           'Benzersiz Kullanıcı': queryset.values('user').distinct().count(),
           'En Çok Kullanılan SAP Modülü': queryset.values('sap_module_detected').annotate(Count('id')).order_by('-id__count').first().get('sap_module_detected', 'N/A') if queryset.exists() else 'N/A',
           'Hata Oranı (%)': round((queryset.filter(error_occurred=True).count() / queryset.count() * 100) if queryset.count() > 0 else 0, 2)
       }
   
   def _format_excel_sheets(self, writer):
       """Excel sheet'lerini formatla"""
       for sheet_name in writer.sheets:
           worksheet = writer.sheets[sheet_name]
           
           # Header formatting
           header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
           header_font = Font(color='FFFFFF', bold=True)
           
           for cell in worksheet[1]:
               cell.fill = header_fill
               cell.font = header_font
               cell.alignment = Alignment(horizontal='center')
           
           # Auto-width
           for column in worksheet.columns:
               max_length = 0
               column_letter = column[0].column_letter
               for cell in column:
                   try:
                       if len(str(cell.value)) > max_length:
                           max_length = len(str(cell.value))
                   except:
                       pass
               adjusted_width = min(max_length + 2, 50)
               worksheet.column_dimensions[column_letter].width = adjusted_width
   
   def _generate_comprehensive_summary(self, request: ExportRequest) -> Dict[str, Any]:
       """Kapsamlı özet oluştur"""
       start_date = request.start_date or (timezone.now() - timedelta(days=30))
       end_date = request.end_date or timezone.now()
       
       # Query analytics
       queries = QueryAnalytics.objects.filter(
           created_at__gte=start_date,
           created_at__lte=end_date
       )
       
       # Usage statistics
       usage_stats = UsageStatistics.objects.filter(
           date__gte=start_date.date(),
           date__lte=end_date.date()
       )
       
       # User feedback
       feedback = UserFeedback.objects.filter(
           created_at__gte=start_date,
           created_at__lte=end_date
       )
       
       # Documents
       documents = DocumentSource.objects.filter(
           created_at__gte=start_date,
           created_at__lte=end_date
       )
       
       return {
           'Rapor Dönemi': f"{start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}",
           'Toplam Sorgu': queries.count(),
           'Başarılı Sorgu': queries.filter(response_generated=True).count(),
           'Benzersiz Kullanıcı': queries.values('user').distinct().count(),
           'Ortalama Yanıt Süresi (sn)': round(queries.aggregate(Avg('response_time'))['response_time__avg'] or 0, 2),
           'Toplam Maliyet (₺)': queries.aggregate(Sum('cost_estimate'))['cost_estimate__sum'] or 0,
           'Yeni Döküman': documents.count(),
           'İşlenen Döküman': documents.filter(processing_status='completed').count(),
           'Toplam Feedback': feedback.count(),
           'Ortalama Puan': round(feedback.aggregate(Avg('rating'))['rating__avg'] or 0, 2),
           'Faydalı Bulunan (%)': round((feedback.filter(is_helpful=True).count() / feedback.count() * 100) if feedback.count() > 0 else 0, 2),
           'Hata Oranı (%)': round((queries.filter(error_occurred=True).count() / queries.count() * 100) if queries.count() > 0 else 0, 2)
       }
   
   def _get_user_performance_data(self, request: ExportRequest) -> List[Dict]:
       """Kullanıcı performans verileri"""
       start_date = request.start_date or (timezone.now() - timedelta(days=30))
       end_date = request.end_date or timezone.now()
       
       user_stats = QueryAnalytics.objects.filter(
           created_at__gte=start_date,
           created_at__lte=end_date,
           user__isnull=False
       ).values(
           'user__email'
       ).annotate(
           total_queries=Count('id'),
           successful_queries=Count('id', filter=Q(response_generated=True)),
           avg_response_time=Avg('response_time'),
           avg_confidence=Avg('confidence_score'),
           total_tokens=Sum('tokens_used'),
           total_cost=Sum('cost_estimate')
       ).order_by('-total_queries')
       
       return [
           {
               'Kullanıcı Email': stat['user__email'],
               'Toplam Sorgu': stat['total_queries'],
               'Başarılı Sorgu': stat['successful_queries'],
               'Başarı Oranı (%)': round((stat['successful_queries'] / stat['total_queries'] * 100), 2),
               'Ort. Yanıt Süresi (sn)': round(stat['avg_response_time'] or 0, 2),
               'Ort. Güven Skoru': round(stat['avg_confidence'] or 0, 3),
               'Toplam Token': stat['total_tokens'] or 0,
               'Toplam Maliyet (₺)': stat['total_cost'] or 0
           }
           for stat in user_stats[:50]  # Top 50 kullanıcı
       ]
   
   def _get_module_usage_data(self, request: ExportRequest) -> List[Dict]:
       """SAP modül kullanım verileri"""
       start_date = request.start_date or (timezone.now() - timedelta(days=30))
       end_date = request.end_date or timezone.now()
       
       module_stats = QueryAnalytics.objects.filter(
           created_at__gte=start_date,
           created_at__lte=end_date,
           sap_module_detected__isnull=False
       ).values(
           'sap_module_detected'
       ).annotate(
           usage_count=Count('id'),
           avg_response_time=Avg('response_time'),
           success_rate=Avg('response_generated'),
           avg_confidence=Avg('confidence_score')
       ).order_by('-usage_count')
       
       return [
           {
               'SAP Modülü': stat['sap_module_detected'],
               'Kullanım Sayısı': stat['usage_count'],
               'Ort. Yanıt Süresi (sn)': round(stat['avg_response_time'] or 0, 2),
               'Başarı Oranı (%)': round((stat['success_rate'] or 0) * 100, 2),
               'Ort. Güven Skoru': round(stat['avg_confidence'] or 0, 3)
           }
           for stat in module_stats
       ]
   
   def _get_error_analysis_data(self, request: ExportRequest) -> List[Dict]:
       """Hata analiz verileri"""
       start_date = request.start_date or (timezone.now() - timedelta(days=30))
       end_date = request.end_date or timezone.now()
       
       error_logs = ErrorLog.objects.filter(
           created_at__gte=start_date,
           created_at__lte=end_date
       ).values(
           'error_type', 'error_level'
       ).annotate(
           count=Count('id')
       ).order_by('-count')
       
       return [
           {
               'Hata Tipi': log['error_type'],
               'Hata Seviyesi': log['error_level'],
               'Oluşma Sayısı': log['count']
           }
           for log in error_logs
       ]
   
   def _get_system_performance_data(self, request: ExportRequest) -> List[Dict]:
       """Sistem performans verileri"""
       start_date = request.start_date or (timezone.now() - timedelta(days=30))
       end_date = request.end_date or timezone.now()
       
       # Günlük performans metrikleri
       daily_performance = QueryAnalytics.objects.filter(
           created_at__gte=start_date,
           created_at__lte=end_date
       ).extra(
           select={'day': 'date(created_at)'}
       ).values('day').annotate(
           total_queries=Count('id'),
           avg_response_time=Avg('response_time'),
           success_rate=Avg('response_generated'),
           total_tokens=Sum('tokens_used')
       ).order_by('day')
       
       return [
           {
               'Tarih': perf['day'],
               'Toplam Sorgu': perf['total_queries'],
               'Ort. Yanıt Süresi (sn)': round(perf['avg_response_time'] or 0, 2),
               'Başarı Oranı (%)': round((perf['success_rate'] or 0) * 100, 2),
               'Toplam Token': perf['total_tokens'] or 0
           }
           for perf in daily_performance
       ]


class ExportService:
   """Ana export servis sınıfı"""
   
   def __init__(self):
       self.chat_exporter = ChatExporter()
       self.document_exporter = DocumentExporter()
       self.analytics_exporter = AnalyticsExporter()
       self.cache = CacheManager()
   
   def export_data(self, export_request: ExportRequest) -> ExportResult:
       """Ana export fonksiyonu"""
       try:
           # Cache kontrolü
           cache_key = self._generate_cache_key(export_request)
           cached_result = self.cache.get(cache_key)
           
           if cached_result:
               logger.info(f"Cache'den export döndürülüyor: {cache_key}")
               return ExportResult(**cached_result)
           
           # Export tipine göre yönlendir
           if export_request.export_type in ['conversations', 'messages']:
               result = self.chat_exporter.export_conversations(export_request)
           elif export_request.export_type == 'documents':
               result = self.document_exporter.export_documents(export_request)
           elif export_request.export_type in ['query_analytics', 'usage_statistics', 'user_feedback', 'comprehensive']:
               result = self.analytics_exporter.export_analytics(export_request)
           else:
               raise ValidationException(f"Desteklenmeyen export tipi: {export_request.export_type}")
           
           # Sonucu cache'le (1 saat)
           self.cache.set(cache_key, asdict(result), 3600)
           
           # Export log'u
           self._log_export_activity(export_request, result)
           
           return result
           
       except Exception as e:
           logger.error(f"Export hatası: {e}")
           raise ExportException(f"Export işlemi başarısız: {str(e)}")
   
   def get_export_status(self, export_id: str) -> Dict[str, Any]:
       """Export durumunu kontrol et"""
       cache_key = f"export_status:{export_id}"
       status = self.cache.get(cache_key)
       
       if not status:
           return {
               'status': 'not_found',
               'message': 'Export bulunamadı'
           }
       
       return status
   
   def schedule_export(self, export_request: ExportRequest, user_id: str = None) -> str:
       """Export'u arka planda planla"""
       import uuid
       from ..tasks import process_export_task
       
       export_id = str(uuid.uuid4())
       
       # Export durumunu cache'le
       status = {
           'export_id': export_id,
           'status': 'queued',
           'created_at': timezone.now().isoformat(),
           'progress': 0,
           'estimated_time': self._estimate_export_time(export_request)
       }
       
       self.cache.set(f"export_status:{export_id}", status, 86400)  # 24 saat
       
       # Celery task başlat
       process_export_task.delay(export_id, asdict(export_request), user_id)
       
       return export_id
   
   def cleanup_expired_exports(self) -> int:
       """Süresi dolmuş export dosyalarını temizle"""
       try:
           # 7 günden eski export dosyalarını bul ve sil
           cutoff_date = timezone.now() - timedelta(days=7)
           
           # Storage'daki exports klasörünü tara
           export_prefix = 'exports/'
           deleted_count = 0
           
           try:
               # Django storage ile dosya listesi al
               if hasattr(default_storage, 'listdir'):
                   directories, files = default_storage.listdir(export_prefix)
                   
                   for file_path in files:
                       full_path = f"{export_prefix}{file_path}"
                       try:
                           # Dosya tarihini kontrol et
                           file_modified = default_storage.get_modified_time(full_path)
                           if file_modified < cutoff_date:
                               default_storage.delete(full_path)
                               deleted_count += 1
                               logger.info(f"Eski export dosyası silindi: {full_path}")
                       except Exception as e:
                           logger.warning(f"Dosya silinirken hata: {full_path} - {e}")
               
           except Exception as e:
               logger.error(f"Export temizlik hatası: {e}")
           
           return deleted_count
           
       except Exception as e:
           logger.error(f"Export cleanup hatası: {e}")
           return 0
   
   def get_export_history(self, user_id: str = None, limit: int = 20) -> List[Dict]:
       """Export geçmişini al"""
       try:
           # Cache'den export geçmişini al
           cache_key = f"export_history:{user_id or 'all'}"
           history = self.cache.get(cache_key)
           
           if not history:
               # Sistem loglarından export aktivitelerini al
               logs = SystemLog.objects.filter(
                   module='export_service',
                   user_id=user_id if user_id else None
               ).order_by('-created_at')[:limit]
               
               history = []
               for log in logs:
                   try:
                       extra_data = log.extra_data
                       history.append({
                           'export_id': extra_data.get('export_id'),
                           'export_type': extra_data.get('export_type'),
                           'format': extra_data.get('format'),
                           'record_count': extra_data.get('record_count'),
                           'file_size': extra_data.get('file_size'),
                           'created_at': log.created_at.isoformat(),
                           'status': 'completed'
                       })
                   except:
                       continue
               
               # 1 saat cache'le
               self.cache.set(cache_key, history, 3600)
           
           return history
           
       except Exception as e:
           logger.error(f"Export history hatası: {e}")
           return []
   
   def _generate_cache_key(self, export_request: ExportRequest) -> str:
       """Export request için cache key oluştur"""
       import hashlib
       
       # Request'i string'e çevir ve hash'le
       request_str = json.dumps(asdict(export_request), sort_keys=True, default=str)
       request_hash = hashlib.md5(request_str.encode()).hexdigest()
       
       return f"export_result:{request_hash}"
   
   def _estimate_export_time(self, export_request: ExportRequest) -> int:
       """Export süresi tahmini (saniye)"""
       base_time = 30  # Temel süre
       
       # Export tipine göre süre ayarla
       type_multipliers = {
           'conversations': 1.0,
           'documents': 2.0,
           'query_analytics': 1.5,
           'usage_statistics': 1.2,
           'comprehensive': 3.0
       }
       
       multiplier = type_multipliers.get(export_request.export_type, 1.0)
       
       # Format'a göre süre ayarla
       if export_request.format == 'pdf':
           multiplier *= 1.5
       elif export_request.format == 'zip':
           multiplier *= 2.0
       
       return int(base_time * multiplier)
   
   def _log_export_activity(self, export_request: ExportRequest, result: ExportResult):
       """Export aktivitesini logla"""
       try:
           SystemLog.objects.create(
               level='INFO',
               message=f"Export tamamlandı: {export_request.export_type}",
               module='export_service',
               function='export_data',
               user_id=export_request.user_id,
               extra_data={
                   'export_type': export_request.export_type,
                   'format': export_request.format,
                   'record_count': result.record_count,
                   'file_size': result.file_size,
                   'file_path': result.file_path,
                   'filters': export_request.filters
               }
           )
       except Exception as e:
           logger.warning(f"Export log hatası: {e}")


class StreamingExporter:
   """Büyük veri setleri için streaming export"""
   
   def __init__(self):
       self.chunk_size = 1000  # Her seferde işlenecek kayıt sayısı
   
   def stream_csv_export(self, queryset: QuerySet, headers: List[str], 
                        row_generator_func) -> StreamingHttpResponse:
       """CSV streaming export"""
       def generate_csv():
           output = io.StringIO()
           writer = csv.writer(output)
           
           # Headers
           writer.writerow(headers)
           yield output.getvalue()
           output.seek(0)
           output.truncate(0)
           
           # Data chunks
           total_count = queryset.count()
           processed = 0
           
           for chunk_start in range(0, total_count, self.chunk_size):
               chunk = queryset[chunk_start:chunk_start + self.chunk_size]
               
               for item in chunk:
                   row = row_generator_func(item)
                   writer.writerow(row)
                   processed += 1
               
               yield output.getvalue()
               output.seek(0)
               output.truncate(0)
               
               # Progress log
               if processed % (self.chunk_size * 5) == 0:
                   logger.info(f"Streaming export progress: {processed}/{total_count}")
       
       response = StreamingHttpResponse(
           generate_csv(),
           content_type='text/csv; charset=utf-8-sig'
       )
       
       timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
       filename = f"sapbot_streaming_export_{timestamp}.csv"
       response['Content-Disposition'] = f'attachment; filename="{filename}"'
       
       return response
   
   def stream_json_export(self, queryset: QuerySet, 
                         object_serializer_func) -> StreamingHttpResponse:
       """JSON streaming export"""
       def generate_json():
           yield '{"data": ['
           
           total_count = queryset.count()
           processed = 0
           first_item = True
           
           for chunk_start in range(0, total_count, self.chunk_size):
               chunk = queryset[chunk_start:chunk_start + self.chunk_size]
               
               for item in chunk:
                   if not first_item:
                       yield ','
                   
                   item_json = json.dumps(
                       object_serializer_func(item), 
                       ensure_ascii=False,
                       default=str
                   )
                   yield item_json
                   
                   first_item = False
                   processed += 1
               
               # Progress log
               if processed % (self.chunk_size * 5) == 0:
                   logger.info(f"JSON streaming export progress: {processed}/{total_count}")
           
           yield ']}'
       
       response = StreamingHttpResponse(
           generate_json(),
           content_type='application/json; charset=utf-8'
       )
       
       timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
       filename = f"sapbot_streaming_export_{timestamp}.json"
       response['Content-Disposition'] = f'attachment; filename="{filename}"'
       
       return response


# Utility functions
def create_export_request(export_type: str, format: str, **kwargs) -> ExportRequest:
   """Export request helper"""
   return ExportRequest(
       export_type=export_type,
       format=format,
       **kwargs
   )


def export_chat_conversations(start_date: datetime = None, end_date: datetime = None,
                           format: str = 'excel', user_id: str = None,
                           **filters) -> ExportResult:
   """Chat konuşmalarını export et - shortcut"""
   service = ExportService()
   request = ExportRequest(
       export_type='conversations',
       format=format,
       start_date=start_date,
       end_date=end_date,
       user_id=user_id,
       filters=filters
   )
   return service.export_data(request)


def export_documents(document_type: str = None, language: str = None,
                   format: str = 'excel', **kwargs) -> ExportResult:
   """Dökümanları export et - shortcut"""
   service = ExportService()
   filters = {
       'document_type': document_type,
       'language': language,
       **kwargs
   }
   request = ExportRequest(
       export_type='documents',
       format=format,
       filters=filters
   )
   return service.export_data(request)


def export_analytics(analytics_type: str = 'query_analytics', 
                   start_date: datetime = None, end_date: datetime = None,
                   format: str = 'excel', **filters) -> ExportResult:
   """Analitik verileri export et - shortcut"""
   service = ExportService()
   request = ExportRequest(
       export_type=analytics_type,
       format=format,
       start_date=start_date,
       end_date=end_date,
       filters=filters
   )
   return service.export_data(request)


def cleanup_old_exports() -> int:
   """Eski export dosyalarını temizle - Celery task için"""
   service = ExportService()
   return service.cleanup_expired_exports()


# Scheduled cleanup task
def schedule_export_cleanup():
   """Export temizlik task'ını planla"""
   from ..tasks import cleanup_exports_task
   cleanup_exports_task.delay()


# Export format validators
SUPPORTED_FORMATS = {
   'conversations': ['csv', 'excel', 'json', 'pdf'],
   'documents': ['csv', 'excel', 'zip'],
   'query_analytics': ['csv', 'excel'],
   'usage_statistics': ['csv', 'excel'],
   'user_feedback': ['csv', 'excel'],
   'comprehensive': ['excel', 'pdf']
}


def validate_export_format(export_type: str, format: str) -> bool:
   """Export format doğrulama"""
   return format.lower() in SUPPORTED_FORMATS.get(export_type, [])


def get_supported_formats(export_type: str) -> List[str]:
   """Desteklenen formatları al"""
   return SUPPORTED_FORMATS.get(export_type, [])


# Export quota management
class ExportQuotaManager:
   """Export kota yöneticisi"""
   
   def __init__(self):
       self.cache = CacheManager()
   
   def check_quota(self, user_id: str, export_type: str) -> Dict[str, Any]:
       """Kullanıcı export kotasını kontrol et"""
       # Günlük limit
       daily_limit = self._get_daily_limit(export_type)
       today = timezone.now().date()
       
       cache_key = f"export_quota:{user_id}:{today}"
       current_usage = self.cache.get(cache_key) or 0
       
       remaining = max(0, daily_limit - current_usage)
       
       return {
           'allowed': remaining > 0,
           'daily_limit': daily_limit,
           'current_usage': current_usage,
           'remaining': remaining,
           'reset_time': (timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)).isoformat()
       }
   
   def increment_usage(self, user_id: str, export_type: str):
       """Kullanıcı export kullanımını artır"""
       today = timezone.now().date()
       cache_key = f"export_quota:{user_id}:{today}"
       
       current_usage = self.cache.get(cache_key) or 0
       self.cache.set(cache_key, current_usage + 1, 86400)  # 24 saat
   
   def _get_daily_limit(self, export_type: str) -> int:
       """Export tipine göre günlük limit"""
       limits = {
           'conversations': 10,
           'documents': 5,
           'query_analytics': 20,
           'usage_statistics': 15,
           'user_feedback': 10,
           'comprehensive': 3
       }
       return limits.get(export_type, 5)
        

